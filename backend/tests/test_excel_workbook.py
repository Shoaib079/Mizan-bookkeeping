"""Shared Excel workbook helpers — titles, headers, money format."""

from __future__ import annotations

from openpyxl import Workbook

from app.core.excel.workbook import (
    MONEY_FORMAT,
    MONEY_FORMAT_ACCOUNTING,
    add_sheet,
    create_workbook,
    finish_data_table,
    money_header,
    quantity_header,
    unique_sheet_title,
    write_header_row,
    write_money,
    write_sheet_title,
)


def test_unique_sheet_title_caps_length_and_avoids_collisions():
    wb = Workbook()
    wb.active.title = "Cash — Short"
    long = "Cash — " + ("VeryLongTurkishRestaurantAccountName" * 3)
    first = unique_sheet_title(wb, long)
    assert len(first) <= 31
    wb.create_sheet(title=first)
    second = unique_sheet_title(wb, long)
    assert second != first
    assert len(second) <= 31
    assert second.endswith("(2)") or " (2)" in second


def test_add_sheet_never_duplicates_titles():
    wb, _ = create_workbook("Bank — Garanti")
    a = add_sheet(wb, "Bank — Garanti")
    b = add_sheet(wb, "Bank — Garanti")
    assert a.title != b.title
    assert len(a.title) <= 31
    assert len(b.title) <= 31


def test_quantity_header_puts_currency_in_the_label():
    assert quantity_header("USD") == "Amount held (USD)"
    assert money_header() == "Amount (₺)"


def test_finish_data_table_freezes_and_filters():
    wb, ws = create_workbook("Demo")
    write_sheet_title(ws, "Demo report", subtitles=["Period: June"], end_col=2)
    header_row = 4
    data_start = write_header_row(ws, header_row, ["Label", money_header()])
    write_money(ws, data_start, 2, 12_345)
    ws.cell(row=data_start, column=1, value="Sample")
    finish_data_table(
        ws, header_row=header_row, last_data_row=data_start, end_col=2
    )
    assert ws.freeze_panes == "A5"
    assert ws.auto_filter.ref is not None
    # Without money_cols the plain money format is left alone.
    assert ws.cell(row=data_start, column=2).number_format == MONEY_FORMAT
    assert ws.cell(row=data_start, column=2).value == 123.45


def test_money_columns_get_accounting_format():
    """Declared money columns show negatives red in parentheses, still numeric."""
    wb, ws = create_workbook("Demo")
    header_row = 1
    data_start = write_header_row(ws, header_row, ["Label", money_header()])
    ws.cell(row=data_start, column=1, value="Refund")
    write_money(ws, data_start, 2, -12_345)
    finish_data_table(
        ws,
        header_row=header_row,
        last_data_row=data_start,
        end_col=2,
        money_cols=(2,),
    )
    cell = ws.cell(row=data_start, column=2)
    assert cell.number_format == MONEY_FORMAT_ACCOUNTING
    assert cell.value == -123.45  # a real number — SUM() still works


def test_finish_data_table_sets_print_layout():
    """Ctrl-P from Excel should fit one page wide and repeat the header row."""
    wb, ws = create_workbook("Demo")
    header_row = 1
    data_start = write_header_row(ws, header_row, ["Label", money_header()])
    ws.cell(row=data_start, column=1, value="Sample")
    write_money(ws, data_start, 2, 5_000)
    finish_data_table(
        ws,
        header_row=header_row,
        last_data_row=data_start,
        end_col=2,
        print_footer="Kebapci Halil · Demo",
    )
    assert ws.page_setup.fitToWidth == 1
    assert ws.print_title_rows in ("1:1", "$1:$1")
    assert ws.oddFooter.left.text == "Kebapci Halil · Demo"
    assert "&P" in (ws.oddFooter.right.text or "")
