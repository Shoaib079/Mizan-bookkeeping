"""S10 — downloaded Excel key totals must match the on-screen report JSON.

Same parameters and the same ``view``. Load the workbook with openpyxl and
compare cell values to the screen response — never grep builders.
S9 already pins cash-book closing, expenses page total, and GL line count;
this module covers the remaining statement surfaces plus one S8 ledger.
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from app.core.auth.types import EntityRole
from app.core.cash.posting import post_cash_movement
from app.core.chart_of_accounts.default_chart import SALES_REVENUE_CODE
from app.core.chart_of_accounts.models import Account
from app.core.chart_of_accounts.seed import seed_default_chart
from app.core.chart_of_accounts.types import AccountNormalBalance
from app.core.ledger.models import JournalEntrySource
from app.core.ledger.posting import PostingLine, post_journal_entry, void_journal_entry
from app.core.partners import profit_allocation as pa
from app.core.period_locks.models import PeriodLockKind
from app.core.period_locks.service import close_period
from app.core.pos import posting as pos_posting
from app.db.session import entity_context
from app.features.auth import service as auth_service
from app.features.auth.schema import MembershipCreate, UserCreate
from app.features.banking import service as banking_service
from app.features.banking.models import MoneyAccountKind
from app.features.banking.schema import MoneyAccountCreate
from app.features.cash.models import CashMovementDirection
from app.features.partners.models import Partner
from app.features.reports import financial_statements
from tests.delivery_helpers import ACTOR_ID
from tests.test_delivery_sales_report import _create_and_post
from tests.test_financial_statements import (
    PERIOD_END,
    PERIOD_START,
    _post_period_sales,
    _post_rent_expense,
)
from tests.test_kdv_input_report import (
    _post_supplier_draft,
    _supplier,
    _supplier_draft,
)
from tests.delivery_helpers import delivery_setup as build_delivery_setup

CASH_CODE = "1000"
JUNE_START = date(2026, 6, 1)
JUNE_END = date(2026, 6, 30)
UNLOCK = "Correcting a duplicate found later"
SALE = 100_000


@pytest.fixture
def parity_setup(db_session, restaurant_a):
    seed_default_chart(db_session, restaurant_a.id)
    drawer = banking_service.create_money_account(
        db_session,
        restaurant_a.id,
        MoneyAccountCreate(account_kind=MoneyAccountKind.CASH, name="Main Drawer"),
    )
    with entity_context(db_session, restaurant_a.id):
        accounts = {a.code: a.id for a in db_session.scalars(select(Account))}
    return {
        "entity_id": restaurant_a.id,
        "drawer": drawer,
        "accounts": accounts,
    }


@pytest.fixture
def sealed_books(db_session, restaurant_a):
    seed_default_chart(db_session, restaurant_a.id)
    owner = auth_service.create_user(
        db_session,
        UserCreate(email="s10-parity-owner@example.com", display_name="Owner"),
    )
    auth_service.add_entity_member(
        db_session,
        restaurant_a.id,
        MembershipCreate(user_id=owner.id, role=EntityRole.OWNER),
    )
    with entity_context(db_session, restaurant_a.id):
        accounts = {a.code: a.id for a in db_session.scalars(select(Account))}
    return {"entity_id": restaurant_a.id, "owner_id": owner.id, "accounts": accounts}


def _load(data: bytes):
    return load_workbook(BytesIO(data)).active


def _lira(kurus: int) -> float:
    return kurus / 100


def _money_beside_label(ws, label: str, money_col: int):
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == label:
            return ws.cell(row=r, column=money_col).value
    raise AssertionError(f"label {label!r} not found")


def _assert_lira(cell_value, kurus: int) -> None:
    assert isinstance(cell_value, (int, float))
    assert cell_value == pytest.approx(_lira(kurus), abs=0.01)


# ---------------------------------------------------------------------------
# Phase 7 statement reports
# ---------------------------------------------------------------------------


def test_profit_and_loss_screen_matches_export(
    db_session, client: TestClient, parity_setup
) -> None:
    setup = parity_setup
    _post_period_sales(db_session, setup)
    _post_rent_expense(
        db_session, setup, amount_kurus=20_000, expense_date=date(2026, 1, 16)
    )
    params = {"from": str(PERIOD_START), "to": str(PERIOD_END)}

    screen = client.get(
        f"/entities/{setup['entity_id']}/reports/profit-and-loss", params=params
    )
    assert screen.status_code == 200
    net = screen.json()["net_income_kurus"]

    export = client.get(
        f"/entities/{setup['entity_id']}/reports/profit-and-loss/export",
        params=params,
    )
    assert export.status_code == 200
    _assert_lira(_money_beside_label(_load(export.content), "NET INCOME", 4), net)


def test_balance_sheet_screen_matches_export(
    db_session, client: TestClient, parity_setup
) -> None:
    setup = parity_setup
    _post_period_sales(db_session, setup)
    params = {"as_of": str(PERIOD_END)}

    screen = client.get(
        f"/entities/{setup['entity_id']}/reports/balance-sheet", params=params
    )
    assert screen.status_code == 200
    body = screen.json()
    assert body["accounting_equation_balanced"] is True

    export = client.get(
        f"/entities/{setup['entity_id']}/reports/balance-sheet/export",
        params=params,
    )
    assert export.status_code == 200
    ws = _load(export.content)
    _assert_lira(
        _money_beside_label(ws, "Total assets", 4), body["total_assets_kurus"]
    )
    _assert_lira(
        _money_beside_label(ws, "Total liabilities and equity", 4),
        body["total_liabilities_and_equity_kurus"],
    )
    eq = _money_beside_label(ws, "Accounting equation balanced", 2)
    assert eq is True or eq == "True"


def test_cash_flow_closing_screen_matches_export(
    db_session, client: TestClient, parity_setup
) -> None:
    setup = parity_setup
    _post_period_sales(db_session, setup)
    params = {"from": str(PERIOD_START), "to": str(PERIOD_END)}

    screen = client.get(
        f"/entities/{setup['entity_id']}/reports/cash-flow", params=params
    )
    assert screen.status_code == 200
    closing = screen.json()["closing_cash_kurus"]

    export = client.get(
        f"/entities/{setup['entity_id']}/reports/cash-flow/export",
        params=params,
    )
    assert export.status_code == 200
    _assert_lira(
        _money_beside_label(_load(export.content), "Closing cash", 2), closing
    )


def test_kdv_total_screen_matches_export(
    db_session, client: TestClient, restaurant_a, parity_setup
) -> None:
    setup = parity_setup
    supplier_id = _supplier(db_session, restaurant_a)
    expense_id = setup["accounts"]["5200"]
    draft = _supplier_draft(
        db_session,
        setup["entity_id"],
        supplier_id,
        invoice_date=date(2026, 5, 1),
        invoice_number="PARITY-KDV",
        net_kurus=200_000,
        gross_kurus=240_000,
        vat_breakdown=[
            {"rate_percent": 20, "base_kurus": 200_000, "vat_kurus": 40_000},
        ],
        file_fingerprint="parity-kdv-fp",
    )
    _post_supplier_draft(db_session, setup["entity_id"], draft.id, expense_id)
    params = {"from": "2026-05-01", "to": "2026-05-31"}

    screen = client.get(
        f"/entities/{setup['entity_id']}/reports/kdv-input", params=params
    )
    assert screen.status_code == 200
    body = screen.json()

    export = client.get(
        f"/entities/{setup['entity_id']}/reports/kdv-input/export",
        params=params,
    )
    assert export.status_code == 200
    ws = _load(export.content)
    _assert_lira(_money_beside_label(ws, "TOTAL", 2), body["total_base_kurus"])
    _assert_lira(_money_beside_label(ws, "TOTAL", 3), body["total_vat_kurus"])


def test_delivery_sales_gross_screen_matches_export(
    db_session, client: TestClient, restaurant_a
) -> None:
    setup = build_delivery_setup(
        db_session, restaurant_a.id, platform_names=("Getir",)
    )
    entity_id = setup["entity_id"]
    _create_and_post(
        db_session, entity_id, setup["platforms"]["Getir"].id, date(2026, 1, 15), 180_000
    )
    params = {"from": "2026-01-01", "to": "2026-01-31"}

    screen = client.get(
        f"/entities/{entity_id}/reports/delivery-sales", params=params
    )
    assert screen.status_code == 200
    total = screen.json()["total_gross_kurus"]
    assert total == 180_000

    export = client.get(
        f"/entities/{entity_id}/reports/delivery-sales/export", params=params
    )
    assert export.status_code == 200
    _assert_lira(_money_beside_label(_load(export.content), "TOTAL", 3), total)


def test_period_comparison_current_metric_screen_matches_export(
    db_session, client: TestClient, parity_setup
) -> None:
    setup = parity_setup
    post_cash_movement(
        db_session,
        setup["entity_id"],
        money_account_id=setup["drawer"].id,
        movement_date=date(2026, 1, 10),
        direction=CashMovementDirection.IN,
        amount_kurus=100_000,
        offset_account_id=setup["accounts"][SALES_REVENUE_CODE],
        description="Cash sales",
        actor_id=ACTOR_ID,
    )
    pos_posting.post_card_sales_batch(
        db_session,
        setup["entity_id"],
        sales_date=date(2026, 1, 12),
        gross_amount_kurus=50_000,
        description="Card sales",
        actor_id=ACTOR_ID,
    )
    params = {"from": "2026-01-01", "to": "2026-01-31"}

    screen = client.get(
        f"/entities/{setup['entity_id']}/reports/period-comparison", params=params
    )
    assert screen.status_code == 200
    metrics = {m["key"]: m for m in screen.json()["metrics"]}
    assert metrics, "expected at least one comparison metric"
    metric = next(iter(metrics.values()))

    export = client.get(
        f"/entities/{setup['entity_id']}/reports/period-comparison/export",
        params=params,
    )
    assert export.status_code == 200
    ws = _load(export.content)
    _assert_lira(
        _money_beside_label(ws, metric["label"], 2), metric["current_kurus"]
    )
    _assert_lira(
        _money_beside_label(ws, metric["label"], 3), metric["prior_kurus"]
    )


# ---------------------------------------------------------------------------
# Same-view discipline (closed month)
# ---------------------------------------------------------------------------


def _june_sale(db_session, books, amount: int = SALE):
    with entity_context(db_session, books["entity_id"]):
        entry = post_journal_entry(
            db_session,
            books["entity_id"],
            date(2026, 6, 10),
            "Cash sale",
            [
                PostingLine(
                    books["accounts"][CASH_CODE], amount, AccountNormalBalance.DEBIT
                ),
                PostingLine(
                    books["accounts"][SALES_REVENUE_CODE],
                    amount,
                    AccountNormalBalance.CREDIT,
                ),
            ],
            actor_id=books["owner_id"],
            source=JournalEntrySource.MANUAL,
        )
        db_session.commit()
        return entry.id


def _void_sale(db_session, books, entry_id):
    with entity_context(db_session, books["entity_id"]):
        void_journal_entry(
            db_session,
            books["entity_id"],
            entry_id,
            actor_id=books["owner_id"],
            reason="Recorded twice",
            void_date=date(2026, 6, 10),
            period_unlock_reason=UNLOCK,
        )
        db_session.commit()


def test_closed_month_pl_sealed_screen_matches_export_and_stamp(
    db_session, client: TestClient, sealed_books
) -> None:
    books = sealed_books
    _june_sale(db_session, books)
    close_period(
        db_session,
        books["entity_id"],
        lock_kind=PeriodLockKind.MONTH,
        anchor_date=JUNE_END,
        actor_id=books["owner_id"],
    )
    params = {
        "from": str(JUNE_START),
        "to": str(JUNE_END),
        "view": financial_statements.VIEW_AS_CLOSED,
    }

    screen = client.get(
        f"/entities/{books['entity_id']}/reports/profit-and-loss", params=params
    )
    assert screen.status_code == 200
    net = screen.json()["net_income_kurus"]
    assert net == SALE

    export = client.get(
        f"/entities/{books['entity_id']}/reports/profit-and-loss/export",
        params=params,
    )
    assert export.status_code == 200
    disposition = export.headers.get("content-disposition", "")
    assert "-as-closed.xlsx" in disposition
    ws = _load(export.content)
    _assert_lira(_money_beside_label(ws, "NET INCOME", 4), net)
    cells = {
        str(ws.cell(row=r, column=c).value)
        for r in range(1, 8)
        for c in range(1, 3)
        if ws.cell(row=r, column=c).value is not None
    }
    assert "As closed" in cells


def test_closed_month_pl_live_screen_matches_export_and_stamp(
    db_session, client: TestClient, sealed_books
) -> None:
    books = sealed_books
    entry_id = _june_sale(db_session, books)
    close_period(
        db_session,
        books["entity_id"],
        lock_kind=PeriodLockKind.MONTH,
        anchor_date=JUNE_END,
        actor_id=books["owner_id"],
    )
    _void_sale(db_session, books, entry_id)
    params = {
        "from": str(JUNE_START),
        "to": str(JUNE_END),
        "view": financial_statements.VIEW_LIVE,
    }

    screen = client.get(
        f"/entities/{books['entity_id']}/reports/profit-and-loss", params=params
    )
    assert screen.status_code == 200
    net = screen.json()["net_income_kurus"]
    assert net == 0

    export = client.get(
        f"/entities/{books['entity_id']}/reports/profit-and-loss/export",
        params=params,
    )
    assert export.status_code == 200
    disposition = export.headers.get("content-disposition", "")
    assert "-live.xlsx" in disposition
    ws = _load(export.content)
    _assert_lira(_money_beside_label(ws, "NET INCOME", 4), net)
    cells = {
        str(ws.cell(row=r, column=c).value)
        for r in range(1, 8)
        for c in range(1, 3)
        if ws.cell(row=r, column=c).value is not None
    }
    assert "Live" in cells


# ---------------------------------------------------------------------------
# S8 surface not pinned by S9
# ---------------------------------------------------------------------------


def test_partner_ledger_running_balance_screen_matches_export(
    db_session, client: TestClient, restaurant_a
) -> None:
    seed_default_chart(db_session, restaurant_a.id)
    with entity_context(db_session, restaurant_a.id):
        partner = Partner(name="Ali", ownership_share_pct=Decimal("100"))
        db_session.add(partner)
        db_session.commit()
        db_session.refresh(partner)
    entity_id = restaurant_a.id
    partner_id = partner.id
    pa.post_profit_allocation(
        db_session,
        entity_id,
        allocation_date=date(2026, 6, 30),
        profit_kurus=100_000,
        description="Parity allocation",
        actor_id=ACTOR_ID,
        net_against_drawings=False,
        netting_as_of=date(2026, 6, 30),
    )

    screen = client.get(f"/entities/{entity_id}/partners/{partner_id}/ledger")
    assert screen.status_code == 200
    body = screen.json()
    # Running column ends on current_account (profit included), not net_balance
    # (settlement helper). Capital is a separate summary sticker.
    current = body["current_account_kurus"]
    capital = body["capital_balance_kurus"]

    export = client.get(
        f"/entities/{entity_id}/partners/{partner_id}/ledger/export"
    )
    assert export.status_code == 200
    ws = _load(export.content)
    # Summary strip: "Net balance" / "Capital in business" in col 2 above Date.
    summary_amounts = [
        ws.cell(row=r, column=2).value
        for r in range(1, ws.max_row + 1)
        if isinstance(ws.cell(row=r, column=2).value, (int, float))
        and ws.cell(row=r, column=1).value
        and "Date" not in str(ws.cell(row=r, column=1).value)
    ]
    running_vals = []
    header = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Date":
            header = r
            break
    assert header is not None
    for r in range(header + 1, ws.max_row + 1):
        cell = ws.cell(row=r, column=5)
        if isinstance(cell.value, (int, float)):
            running_vals.append(cell.value)
    assert running_vals
    _assert_lira(running_vals[-1], current)
    assert any(
        abs(v - _lira(current)) < 0.01 or abs(v - _lira(capital)) < 0.01
        for v in summary_amounts
    )
