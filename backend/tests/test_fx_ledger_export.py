"""FX hub ledger Excel/PDF download endpoints."""

from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.core.chart_of_accounts.seed import seed_default_chart
from app.core.fx import posting as fx_posting
from app.features.banking import service as banking_service
from app.features.banking.models import MoneyAccountKind
from app.features.banking.schema import MoneyAccountCreate


ACTOR_ID = __import__("uuid").UUID("00000000-0000-4000-8000-000000000001")


def _try_cash(db_session, entity_id, name: str = "Main Drawer"):
    return banking_service.create_money_account(
        db_session,
        entity_id,
        MoneyAccountCreate(account_kind=MoneyAccountKind.CASH, name=name),
    )


def _fx_wallet(db_session, entity_id, currency: str = "USD", name: str = "USD Wallet"):
    return banking_service.create_money_account(
        db_session,
        entity_id,
        MoneyAccountCreate(
            account_kind=MoneyAccountKind.FOREIGN_CURRENCY,
            currency=currency,
            name=name,
        ),
    )


@pytest.fixture
def fx_export_setup(db_session, restaurant_a):
    seed_default_chart(db_session, restaurant_a.id)
    drawer = _try_cash(db_session, restaurant_a.id)
    usd_wallet = _fx_wallet(db_session, restaurant_a.id)
    fx_posting.post_fx_purchase(
        db_session,
        restaurant_a.id,
        fx_money_account_id=usd_wallet.id,
        try_cash_money_account_id=drawer.id,
        native_quantity=10_000,
        try_cost_kurus=350_000,
        purchase_date=date(2026, 6, 15),
        description="Export sample purchase",
        actor_id=ACTOR_ID,
    )
    return {
        "entity_id": restaurant_a.id,
        "purchase_date": "2026-06-15",
    }


def test_fx_hub_ledger_export_xlsx_and_pdf(
    fx_export_setup, client: TestClient
) -> None:
    entity_id = fx_export_setup["entity_id"]
    purchase_date = fx_export_setup["purchase_date"]
    params = {"from": purchase_date, "to": purchase_date}

    xlsx = client.get(f"/entities/{entity_id}/fx/ledger/export", params=params)
    assert xlsx.status_code == 200, xlsx.text
    assert "spreadsheetml" in xlsx.headers["content-type"]
    assert xlsx.content[:2] == b"PK"

    pdf = client.get(f"/entities/{entity_id}/fx/ledger/export/pdf", params=params)
    assert pdf.status_code == 200, pdf.text
    assert pdf.headers["content-type"].startswith("application/pdf")
    assert pdf.content[:4] == b"%PDF"


def test_fx_hub_ledger_collect_rows(db_session, fx_export_setup) -> None:
    from datetime import date as date_cls

    from app.features.fx import hub_export as fx_hub_export

    rows, label = fx_hub_export.collect_fx_hub_ledger_rows(
        db_session,
        fx_export_setup["entity_id"],
        from_date=date_cls.fromisoformat(fx_export_setup["purchase_date"]),
        to_date=date_cls.fromisoformat(fx_export_setup["purchase_date"]),
    )
    assert label == "All wallets"
    assert len(rows) == 1
    assert "FX purchase" in rows[0].description
    assert rows[0].wallet_name == "USD Wallet"


def test_fx_hub_ledger_export_contains_movement(
    fx_export_setup, client: TestClient
) -> None:
    entity_id = fx_export_setup["entity_id"]
    purchase_date = fx_export_setup["purchase_date"]
    resp = client.get(
        f"/entities/{entity_id}/fx/ledger/export",
        params={"from": purchase_date, "to": purchase_date},
    )
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.content))
    ws = wb["FX ledger"]
    header = next(
        row[0].row for row in ws.iter_rows(min_col=1, max_col=1) if row[0].value == "Date"
    )
    rows = list(ws.iter_rows(min_row=header + 1, values_only=True))
    flat = [str(cell) for row in rows if row for cell in row if cell is not None]
    assert any(
        "Export sample purchase" in text or "FX purchase" in text for text in flat
    )


def test_fx_hub_ledger_export_wallet_filter(
    db_session, restaurant_a, fx_export_setup, client: TestClient
) -> None:
    entity_id = fx_export_setup["entity_id"]
    purchase_date = fx_export_setup["purchase_date"]
    eur_wallet = _fx_wallet(
        db_session, restaurant_a.id, currency="EUR", name="EUR Wallet"
    )
    drawer = _try_cash(db_session, restaurant_a.id, name="Drawer 2")
    fx_posting.post_fx_purchase(
        db_session,
        restaurant_a.id,
        fx_money_account_id=eur_wallet.id,
        try_cash_money_account_id=drawer.id,
        native_quantity=5_000,
        try_cost_kurus=180_000,
        purchase_date=date(2026, 6, 15),
        description="EUR only row",
        actor_id=ACTOR_ID,
    )

    usd_only = client.get(
        f"/entities/{entity_id}/fx/ledger/export",
        params={"from": purchase_date, "to": purchase_date, "wallet": "USD"},
    )
    assert usd_only.status_code == 200
    wb = load_workbook(BytesIO(usd_only.content))
    ws = wb["FX ledger"]
    flat = [
        str(cell)
        for row in ws.iter_rows(min_row=1, values_only=True)
        if row
        for cell in row
        if cell is not None
    ]
    assert any("USD Wallet" in text for text in flat)
    assert not any("EUR Wallet" in text for text in flat)
