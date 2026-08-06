"""Staff, customer and supplier ledger downloads.

Partner had the only export in the app. These three were missing entirely —
no route, no button — which is a gap that looks like a design decision rather
than a fault, so it went unnoticed for a long time. All four now build from
`reports.subledger_export`; `test_partner_ledger_export.py` covers that the
partner output did not change when it moved onto the shared builder.
"""

from __future__ import annotations

import io
import uuid
from datetime import date, datetime

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.core.chart_of_accounts.seed import seed_default_chart
from app.core.payables.types import SupplierMovementType
from app.core.receivables.ledger import persist_customer_ledger_entry
from app.core.receivables.types import CustomerMovementType
from app.db.session import entity_context
from app.features.customers.models import Customer
from app.core.staff.types import PayCurrency
from app.features.staff.models import Employee
from app.features.suppliers.models import Supplier

ACTOR_ID = uuid.UUID("00000000-0000-4000-8000-000000000001")


def _sheet(content: bytes):
    return load_workbook(io.BytesIO(content)).active


def _column(ws, index: int, *, min_row: int) -> list:
    return [row[index].value for row in ws.iter_rows(min_row=min_row) if row[index].value]


# --------------------------------------------------------------------------
# Staff
# --------------------------------------------------------------------------


def _employee(db_session, entity_id) -> Employee:
    seed_default_chart(db_session, entity_id)
    with entity_context(db_session, entity_id):
        employee = Employee(name="Ayşe Demir", pay_currency=PayCurrency.TRY)
        db_session.add(employee)
        db_session.commit()
        db_session.refresh(employee)
    return employee


def test_staff_ledger_downloads_as_xlsx_and_pdf(
    db_session, restaurant_a, client: TestClient
) -> None:
    employee = _employee(db_session, restaurant_a.id)
    base = f"/entities/{restaurant_a.id}/staff/employees/{employee.id}/ledger"

    xlsx = client.get(f"{base}/export")
    assert xlsx.status_code == 200, xlsx.text
    assert "spreadsheetml" in xlsx.headers["content-type"]
    assert xlsx.content[:2] == b"PK"

    pdf = client.get(f"{base}/export/pdf")
    assert pdf.status_code == 200, pdf.text
    assert pdf.headers["content-type"].startswith("application/pdf")
    assert pdf.content[:4] == b"%PDF"


def test_staff_export_names_the_employee_and_the_restaurant(
    db_session, restaurant_a, client: TestClient
) -> None:
    """A folder of these has to be readable without opening them."""
    employee = _employee(db_session, restaurant_a.id)
    resp = client.get(
        f"/entities/{restaurant_a.id}/staff/employees/{employee.id}/ledger/export"
    )
    assert resp.status_code == 200, resp.text
    disposition = resp.headers["content-disposition"]
    assert "staff" in disposition
    assert "Restaurant" in disposition or "restaurant" in disposition.lower()


def test_staff_export_heads_the_sheet_with_the_salary_figures(
    db_session, restaurant_a, client: TestClient
) -> None:
    employee = _employee(db_session, restaurant_a.id)
    resp = client.get(
        f"/entities/{restaurant_a.id}/staff/employees/{employee.id}/ledger/export"
    )
    ws = _sheet(resp.content)
    labels = [ws.cell(row=r, column=1).value for r in range(3, 6)]
    assert labels == ["Balance", "Unpaid salary", "Outstanding advance"], labels


def test_staff_export_404s_for_an_unknown_employee(
    db_session, restaurant_a, client: TestClient
) -> None:
    seed_default_chart(db_session, restaurant_a.id)
    resp = client.get(
        f"/entities/{restaurant_a.id}/staff/employees/{uuid.uuid4()}/ledger/export"
    )
    assert resp.status_code == 404, resp.text


# --------------------------------------------------------------------------
# Customer
# --------------------------------------------------------------------------


def _customer_with_sale(db_session, entity_id) -> Customer:
    seed_default_chart(db_session, entity_id)
    with entity_context(db_session, entity_id):
        customer = Customer(entity_id=entity_id, name="Blue Tours")
        db_session.add(customer)
        db_session.flush()
        persist_customer_ledger_entry(
            db_session,
            customer.id,
            movement_date=date(2026, 6, 30),
            movement_type=CustomerMovementType.CREDIT_SALE,
            amount_kurus=413_600,
            description="Dinner for 6",
            actor_id=ACTOR_ID,
        )
        db_session.commit()
        db_session.refresh(customer)
    return customer


def test_customer_ledger_downloads_as_xlsx_and_pdf(
    db_session, restaurant_a, client: TestClient
) -> None:
    customer = _customer_with_sale(db_session, restaurant_a.id)
    base = f"/entities/{restaurant_a.id}/customers/{customer.id}/ledger"

    xlsx = client.get(f"{base}/export")
    assert xlsx.status_code == 200, xlsx.text
    assert xlsx.content[:2] == b"PK"

    pdf = client.get(f"{base}/export/pdf")
    assert pdf.status_code == 200, pdf.text
    assert pdf.content[:4] == b"%PDF"


def test_customer_export_writes_movements_in_words_and_real_dates(
    db_session, restaurant_a, client: TestClient
) -> None:
    """"Credit sale", not "credit_sale" — and a date cell, not a string.

    A customer reading their own statement should not have to decode a
    database enum, and a date written as text sorts alphabetically in Excel.
    """
    customer = _customer_with_sale(db_session, restaurant_a.id)
    resp = client.get(
        f"/entities/{restaurant_a.id}/customers/{customer.id}/ledger/export"
    )
    ws = _sheet(resp.content)
    # One summary row (Balance) at row 3, blank row 4, header row 5, data from 6.
    movements = _column(ws, 1, min_row=6)
    assert "Credit sale" in movements, movements
    assert not any("_" in str(m) for m in movements), movements

    dates = _column(ws, 0, min_row=6)
    assert dates, "expected a movement row"
    assert all(isinstance(value, (date, datetime)) for value in dates), dates


def test_customer_export_404s_for_an_unknown_customer(
    db_session, restaurant_a, client: TestClient
) -> None:
    seed_default_chart(db_session, restaurant_a.id)
    resp = client.get(
        f"/entities/{restaurant_a.id}/customers/{uuid.uuid4()}/ledger/export"
    )
    assert resp.status_code == 404, resp.text


# --------------------------------------------------------------------------
# Supplier
# --------------------------------------------------------------------------


def _supplier(db_session, entity_id) -> Supplier:
    seed_default_chart(db_session, entity_id)
    with entity_context(db_session, entity_id):
        # vkn is not nullable — a supplier without a tax number will not insert.
        supplier = Supplier(name="Anadolu Gıda", vkn="2000000001")
        db_session.add(supplier)
        db_session.commit()
        db_session.refresh(supplier)
    return supplier


def test_supplier_ledger_downloads_as_xlsx_and_pdf(
    db_session, restaurant_a, client: TestClient
) -> None:
    supplier = _supplier(db_session, restaurant_a.id)
    base = f"/entities/{restaurant_a.id}/suppliers/{supplier.id}/ledger"

    xlsx = client.get(f"{base}/export")
    assert xlsx.status_code == 200, xlsx.text
    assert xlsx.content[:2] == b"PK"

    pdf = client.get(f"{base}/export/pdf")
    assert pdf.status_code == 200, pdf.text
    assert pdf.content[:4] == b"%PDF"


def test_supplier_export_404s_for_an_unknown_supplier(
    db_session, restaurant_a, client: TestClient
) -> None:
    seed_default_chart(db_session, restaurant_a.id)
    resp = client.get(
        f"/entities/{restaurant_a.id}/suppliers/{uuid.uuid4()}/ledger/export"
    )
    assert resp.status_code == 404, resp.text


def test_supplier_movement_labels_are_words(
    db_session, restaurant_a, client: TestClient
) -> None:
    from app.core.excel.labels import format_supplier_movement

    assert format_supplier_movement(SupplierMovementType.CREDIT_NOTE) == "Credit note"
    assert format_supplier_movement(SupplierMovementType.INVOICE) == "Invoice"
    assert format_supplier_movement(None) == ""
