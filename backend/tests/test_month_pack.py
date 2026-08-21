"""One workbook with every book for the period — the file sent to partners.

Checking a month used to mean six separate downloads, and four of the books
(expenses, cash, bank, ledger) had no export at all.
"""

from __future__ import annotations

import uuid
from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from app.core.auth.types import EntityRole
from app.core.chart_of_accounts.default_chart import SALES_REVENUE_CODE
from app.core.chart_of_accounts.models import Account
from app.core.chart_of_accounts.seed import seed_default_chart
from app.core.chart_of_accounts.types import AccountNormalBalance
from app.core.excel.workbook import MONEY_FORMAT_ACCOUNTING, money_header, quantity_header
from app.core.fx import posting as fx_posting
from app.core.ledger.models import JournalEntrySource
from app.core.ledger.posting import PostingLine, post_journal_entry
from app.core.period_locks.models import PeriodLockKind
from app.core.period_locks.service import close_period
from app.core.staff import posting as staff_posting
from app.core.staff.types import PayCurrency
from app.db.session import entity_context
from app.features.auth import service as auth_service
from app.features.auth.schema import MembershipCreate, UserCreate
from app.features.banking import service as banking_service
from app.features.banking.models import MoneyAccountKind
from app.features.banking.schema import MoneyAccountCreate
from app.features.reports import month_pack
from app.features.staff.models import Employee

CASH_CODE = "1000"
JUNE_START = date(2026, 6, 1)
JUNE_END = date(2026, 6, 30)


@pytest.fixture
def books(db_session, restaurant_a):
    seed_default_chart(db_session, restaurant_a.id)
    owner = auth_service.create_user(
        db_session,
        UserCreate(email="pack-owner@example.com", display_name="Owner"),
    )
    auth_service.add_entity_member(
        db_session,
        restaurant_a.id,
        MembershipCreate(user_id=owner.id, role=EntityRole.OWNER),
    )
    banking_service.create_money_account(
        db_session,
        restaurant_a.id,
        MoneyAccountCreate(account_kind=MoneyAccountKind.CASH, name="Main Drawer"),
    )
    banking_service.create_money_account(
        db_session,
        restaurant_a.id,
        MoneyAccountCreate(
            account_kind=MoneyAccountKind.BANK, name="Garanti", bank_name="Garanti"
        ),
    )
    with entity_context(db_session, restaurant_a.id):
        accounts = {a.code: a.id for a in db_session.scalars(select(Account))}
    return {"entity_id": restaurant_a.id, "owner_id": owner.id, "accounts": accounts}


def _sale(db_session, books, on: date, amount: int):
    with entity_context(db_session, books["entity_id"]):
        post_journal_entry(
            db_session,
            books["entity_id"],
            on,
            "Cash sale",
            [
                PostingLine(
                    books["accounts"][CASH_CODE], amount, AccountNormalBalance.DEBIT
                ),
                PostingLine(
                    books["accounts"][SALES_REVENUE_CODE],
                    amount,
                    AccountNormalBalance.CREDIT,
                ),
            ],
            actor_id=books["owner_id"],
            source=JournalEntrySource.MANUAL,
        )
        db_session.commit()


def _pack(db_session, books, from_date=JUNE_START, to_date=JUNE_END):
    data, ctx = month_pack.build_month_pack_xlsx(
        db_session, books["entity_id"], from_date, to_date
    )
    return load_workbook(BytesIO(data)), ctx


def test_the_pack_holds_every_book(db_session, books):
    _sale(db_session, books, date(2026, 6, 10), 100_000)
    wb, _ = _pack(db_session, books)

    names = wb.sheetnames
    assert "Summary" in names
    assert "Sales" in names
    assert "Expenses" in names
    assert "Salaries" in names
    assert "Card clearing" in names
    assert "Profit and loss" in names
    assert "General ledger" in names
    # One sheet per money account, named so a partner knows which is which.
    assert any(n.startswith("Cash — ") for n in names)
    assert any(n.startswith("Bank — ") for n in names)


def test_the_summary_names_the_period_and_the_business(db_session, books):
    _sale(db_session, books, date(2026, 6, 10), 100_000)
    wb, _ = _pack(db_session, books)

    summary = wb["Summary"]
    assert "01.06.2026 – 30.06.2026" in str(summary.cell(row=2, column=2).value)
    assert "books for the period" in str(summary.cell(row=1, column=1).value)


def test_an_open_month_says_it_is_live(db_session, books):
    """A partner must be able to tell a draft from a sealed month."""
    _sale(db_session, books, date(2026, 6, 10), 100_000)
    wb, ctx = _pack(db_session, books)

    assert ctx.sealed is False
    assert "Live" in str(wb["Summary"].cell(row=3, column=2).value)


def test_a_closed_month_exports_its_sealed_figures(db_session, books):
    """Two partners downloading on different days must get the same file."""
    _sale(db_session, books, date(2026, 6, 10), 100_000)
    close_period(
        db_session,
        books["entity_id"],
        lock_kind=PeriodLockKind.MONTH,
        anchor_date=JUNE_END,
        actor_id=books["owner_id"],
    )

    wb, ctx = _pack(db_session, books)
    assert ctx.sealed is True
    assert "As closed" in str(wb["Summary"].cell(row=3, column=2).value)


def test_the_filename_says_which_it_is(db_session, books):
    _sale(db_session, books, date(2026, 6, 10), 100_000)
    _, ctx = _pack(db_session, books)
    assert month_pack.month_pack_filename(ctx).endswith("-live.xlsx")

    close_period(
        db_session,
        books["entity_id"],
        lock_kind=PeriodLockKind.MONTH,
        anchor_date=JUNE_END,
        actor_id=books["owner_id"],
    )
    _, sealed_ctx = _pack(db_session, books)
    assert month_pack.month_pack_filename(sealed_ctx).endswith("-as-closed.xlsx")


def test_a_quiet_period_still_produces_a_readable_file(db_session, books):
    """Nothing traded — the pack must open, not crash or come out blank."""
    wb, _ = _pack(db_session, books)
    assert "Summary" in wb.sheetnames
    assert "General ledger" in wb.sheetnames


def test_the_ledger_sheet_names_accounts_not_ids(db_session, books):
    """Journal lines carry only an account id; a raw UUID is useless to a
    partner, so names are resolved before writing."""
    _sale(db_session, books, date(2026, 6, 10), 100_000)
    wb, _ = _pack(db_session, books)

    ledger = wb["General ledger"]
    labels = [
        str(ledger.cell(row=r, column=5).value)
        for r in range(4, ledger.max_row + 1)
    ]
    assert any(SALES_REVENUE_CODE in label for label in labels)
    assert not any(label.count("-") == 4 for label in labels if label != "None")


def test_amounts_are_lira_not_kurus(db_session, books):
    """A column of raw kuruş can't be checked against a statement without
    dividing every figure by 100 by hand (2026-07-29)."""
    _sale(db_session, books, date(2026, 6, 10), 123_456)
    wb, _ = _pack(db_session, books)

    summary = wb["Summary"]
    values = [
        summary.cell(row=r, column=2).value for r in range(1, summary.max_row + 1)
    ]
    assert 1234.56 in values, "1.234,56 ₺ should read as 1234.56, not 123456"
    assert 123_456 not in values


def test_money_stays_a_number_so_excel_can_total_it(db_session, books):
    """Formatting as text would look right and break every SUM()."""
    _sale(db_session, books, date(2026, 6, 10), 100_000)
    wb, _ = _pack(db_session, books)

    summary = wb["Summary"]
    money = [
        summary.cell(row=r, column=2)
        for r in range(1, summary.max_row + 1)
        if isinstance(summary.cell(row=r, column=2).value, (int, float))
    ]
    assert money, "expected at least one numeric money cell"
    assert all(isinstance(cell.value, (int, float)) for cell in money)
    assert all(cell.number_format == MONEY_FORMAT_ACCOUNTING for cell in money)


def test_the_pack_shows_foreign_currency_held(db_session, books):
    """It was missing entirely — 'what you hold' wasn't answering the question."""
    wb, _ = _pack(db_session, books)
    assert "Foreign currency" in wb.sheetnames

    fx = wb["Foreign currency"]
    heading = str(fx.cell(row=1, column=1).value)
    assert "Foreign currency held" in heading


def test_the_summary_counts_forex_in_what_you_hold(db_session, books):
    _sale(db_session, books, date(2026, 6, 10), 100_000)
    wb, _ = _pack(db_session, books)

    summary = wb["Summary"]
    labels = [
        str(summary.cell(row=r, column=1).value)
        for r in range(1, summary.max_row + 1)
    ]
    assert any("Foreign currency" in label for label in labels)


def _buy_usd(db_session, books, *, native: int, try_cost: int, on: date):
    wallet = banking_service.create_money_account(
        db_session,
        books["entity_id"],
        MoneyAccountCreate(
            account_kind=MoneyAccountKind.FOREIGN_CURRENCY,
            currency="USD",
            name="USD Wallet",
        ),
    )
    drawers = banking_service.list_money_accounts(
        db_session,
        books["entity_id"],
        account_kind=MoneyAccountKind.CASH,
    )[0]
    drawer = drawers[0]
    fx_posting.post_fx_purchase(
        db_session,
        books["entity_id"],
        fx_money_account_id=wallet.id,
        try_cash_money_account_id=drawer.id,
        native_quantity=native,
        try_cost_kurus=try_cost,
        purchase_date=on,
        description="Buy USD for pack",
        actor_id=books["owner_id"],
    )
    return wallet


def test_foreign_currency_sheet_shows_native_quantity_and_try_cost(db_session, books):
    """Partners need the real USD/EUR held — not only a lira book-cost line."""
    _buy_usd(db_session, books, native=10_000, try_cost=350_000, on=date(2026, 6, 5))
    wb, _ = _pack(db_session, books)

    fx = wb["Foreign currency"]
    natives = [
        fx.cell(row=r, column=3).value for r in range(1, fx.max_row + 1)
    ]
    try_costs = [
        fx.cell(row=r, column=4).value for r in range(1, fx.max_row + 1)
    ]
    assert 100.0 in natives  # $100.00, not 10000 kuruş-style cents left raw
    assert 3500.0 in try_costs  # ₺3.500,00 book cost
    headers = [fx.cell(row=4, column=c).value for c in range(1, 5)]
    assert headers[2] == quantity_header("USD")
    assert headers[3] == money_header("TRY cost")


def test_each_fx_wallet_gets_a_movement_book(db_session, books):
    wallet = _buy_usd(
        db_session, books, native=10_000, try_cost=350_000, on=date(2026, 6, 5)
    )
    wb, _ = _pack(db_session, books)

    sheet_name = next(n for n in wb.sheetnames if n.startswith("FX — "))
    assert wallet.name.split()[0] in sheet_name or "USD" in sheet_name

    book = wb[sheet_name]
    types = [
        str(book.cell(row=r, column=2).value)
        for r in range(1, book.max_row + 1)
    ]
    assert any("purchase" in t.lower() for t in types)
    natives = [
        book.cell(row=r, column=4).value for r in range(1, book.max_row + 1)
    ]
    assert 100.0 in natives


def test_fx_staff_salary_is_not_labelled_as_lira(db_session, books):
    """FX amount_minor is foreign cents — writing it under Amount (₺) lied."""
    _buy_usd(db_session, books, native=200_000, try_cost=7_000_000, on=date(2026, 6, 1))
    with entity_context(db_session, books["entity_id"]):
        employee = Employee(name="FX Cook", pay_currency=PayCurrency.USD)
        db_session.add(employee)
        db_session.commit()
        db_session.refresh(employee)
        employee_id = employee.id

    staff_posting.post_salary_accrual(
        db_session,
        books["entity_id"],
        employee_id,
        accrual_date=date(2026, 6, 10),
        amount_minor=100_000,
        description="USD salary",
        actor_id=books["owner_id"],
        period_year=2026,
        period_month=6,
    )

    wb, _ = _pack(db_session, books)
    salaries = wb["Salaries"]
    headers = [
        salaries.cell(row=5, column=c).value for c in range(1, 8)
    ]
    assert headers[4] == "Currency"
    assert headers[5] == quantity_header("USD", "Amount")
    assert headers[5] != money_header()
    assert headers[6] == money_header("TRY cost")

    currencies = [
        salaries.cell(row=r, column=5).value
        for r in range(6, salaries.max_row + 1)
    ]
    amounts = [
        salaries.cell(row=r, column=6).value
        for r in range(6, salaries.max_row + 1)
    ]
    assert "USD" in currencies
    assert 1000.0 in amounts  # $1,000.00 — must not appear as ₺1.000,00 under a ₺ header


def test_an_unknown_entity_is_a_lookup_error(db_session):
    with pytest.raises(LookupError):
        month_pack.build_month_pack_xlsx(
            db_session, uuid.uuid4(), JUNE_START, JUNE_END
        )


def _pdf_text(data: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(BytesIO(data))
    return "".join(page.extract_text() or "" for page in reader.pages)


def test_summary_cash_rollforward_balances(db_session, books):
    """Opening (day before From) + period cash lines = closing (To)."""
    drawers = banking_service.list_money_accounts(
        db_session,
        books["entity_id"],
        account_kind=MoneyAccountKind.CASH,
    )[0]
    drawer_gl = drawers[0].gl_account_id
    with entity_context(db_session, books["entity_id"]):
        post_journal_entry(
            db_session,
            books["entity_id"],
            date(2026, 6, 10),
            "Cash sale via drawer",
            [
                PostingLine(drawer_gl, 50_000, AccountNormalBalance.DEBIT),
                PostingLine(
                    books["accounts"][SALES_REVENUE_CODE],
                    50_000,
                    AccountNormalBalance.CREDIT,
                ),
            ],
            actor_id=books["owner_id"],
            source=JournalEntrySource.CASH_MOVEMENT,
        )
        db_session.commit()

    bundle = month_pack.load_month_pack_bundle(
        db_session, books["entity_id"], JUNE_START, JUNE_END
    )
    bridge = bundle.cash_bridge
    assert bridge.opening_date == date(2026, 5, 31)
    assert bridge.closing_date == JUNE_END
    assert bridge.closing_cash_bank_kurus == (
        bridge.cash_in_hand_kurus + bridge.bank_balance_kurus
    )
    movement_total = sum(v for _, v in month_pack.cash_movement_rows(bundle.cash_flow))
    assert bridge.balances_with_movements(movement_total)
    assert bridge.balances_with_movements(bundle.cash_flow.net_change_kurus)


def test_excel_summary_shows_rollforward_not_profit_walk(db_session, books):
    drawers = banking_service.list_money_accounts(
        db_session,
        books["entity_id"],
        account_kind=MoneyAccountKind.CASH,
    )[0]
    drawer_gl = drawers[0].gl_account_id
    with entity_context(db_session, books["entity_id"]):
        post_journal_entry(
            db_session,
            books["entity_id"],
            date(2026, 6, 10),
            "Cash sale via drawer",
            [
                PostingLine(drawer_gl, 50_000, AccountNormalBalance.DEBIT),
                PostingLine(
                    books["accounts"][SALES_REVENUE_CODE],
                    50_000,
                    AccountNormalBalance.CREDIT,
                ),
            ],
            actor_id=books["owner_id"],
            source=JournalEntrySource.CASH_MOVEMENT,
        )
        db_session.commit()

    wb, _ = _pack(db_session, books)
    labels = [
        str(wb["Summary"].cell(row=r, column=1).value)
        for r in range(1, wb["Summary"].max_row + 1)
    ]
    assert any("Sales & result" in label for label in labels)
    assert any(label == "Cash & bank" for label in labels)
    # dd.mm.yyyy — these captions used to interpolate the date object straight
    # into the f-string, so they read ISO while every other date on the sheet
    # read 31.05.2026.
    assert any("Opening cash & bank (31.05.2026)" in label for label in labels)
    assert any("Closing cash & bank (30.06.2026)" in label for label in labels)
    assert any(label == "Cash movement" for label in labels)
    assert any("What we hold / owe" in label for label in labels)
    assert any(label == "Cash in hand" for label in labels)
    assert any(label == "Bank" for label in labels)
    assert not any("From net result to cash" in label for label in labels)
    assert not any("CHANGE IN CASH & BANK THIS PERIOD" in label for label in labels)
    assert not any("Other movements" in label for label in labels)


def test_month_pack_pdf_is_a_valid_readable_export(db_session, books):
    _sale(db_session, books, date(2026, 6, 10), 100_000)
    data, ctx = month_pack.build_month_pack_pdf(
        db_session, books["entity_id"], JUNE_START, JUNE_END
    )

    assert data[:4] == b"%PDF"
    assert month_pack.month_pack_pdf_filename(ctx).endswith("-live.pdf")
    text = _pdf_text(data)
    # Dates read the way they do everywhere else in the app (01.06.2026).
    assert "01.06.2026" in text
    assert "30.06.2026" in text
    assert "Summary" in text
    assert "Cash & bank" in text
    assert "Opening cash" in text
    assert "Closing cash" in text
    assert "From net result to cash" not in text
    assert "Expenses" in text
    assert "Profit and loss" in text
    assert "₺" in text


def test_the_pdf_and_the_workbook_state_the_same_period(db_session, books):
    """One download, two files — they must not describe the period differently.

    They drifted once: the PDF moved onto the shared date presentation and the
    workbook kept interpolating raw dates, so the same month pack read
    "01.06.2026 – 30.06.2026" on paper and "2026-06-01 to 2026-06-30" in Excel.
    """
    _sale(db_session, books, date(2026, 6, 10), 100_000)

    wb, _ = _pack(db_session, books)
    workbook_period = str(wb["Summary"].cell(row=2, column=2).value)

    data, _ctx = month_pack.build_month_pack_pdf(
        db_session, books["entity_id"], JUNE_START, JUNE_END
    )
    pdf_text = _pdf_text(data)

    for part in workbook_period.replace("–", " ").split():
        assert part in pdf_text, f"{part!r} is in the workbook but not the PDF"


def test_month_pack_pdf_api(db_session, client, books):
    from app.features.reports import pdf_export

    _sale(db_session, books, date(2026, 6, 10), 100_000)
    response = client.get(
        f"/entities/{books['entity_id']}/reports/month-pack/export/pdf",
        params={"from": "2026-06-01", "to": "2026-06-30"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == pdf_export.PDF_CONTENT_TYPE
    assert 'filename="restaurant-a-books-2026-06-live.pdf"' in response.headers.get(
        "content-disposition", ""
    )
    assert response.content[:4] == b"%PDF"


def test_month_pack_pdf_invalid_date_range(client, books):
    response = client.get(
        f"/entities/{books['entity_id']}/reports/month-pack/export/pdf",
        params={"from": "2026-06-02", "to": "2026-06-01"},
    )
    assert response.status_code == 422


def test_month_pack_api_partial_month_range_in_filename(client, books):
    response = client.get(
        f"/entities/{books['entity_id']}/reports/month-pack",
        params={"from": "2026-08-15", "to": "2026-08-30"},
    )

    assert response.status_code == 200
    disposition = response.headers.get("content-disposition", "")
    assert "2026-08-15-2026-08-30" in disposition


def test_month_pack_api_cross_month_stays_live_when_august_closed(
    db_session, client, books
):
    _sale(db_session, books, date(2026, 8, 20), 100_000)
    close_period(
        db_session,
        books["entity_id"],
        lock_kind=PeriodLockKind.MONTH,
        anchor_date=date(2026, 8, 31),
        actor_id=books["owner_id"],
    )

    response = client.get(
        f"/entities/{books['entity_id']}/reports/month-pack",
        params={"from": "2026-08-15", "to": "2026-09-30"},
    )

    assert response.status_code == 200
    disposition = response.headers.get("content-disposition", "")
    assert "2026-08-15" in disposition
    assert "2026-09-30" in disposition
    assert 'filename="restaurant-a-books-2026-08-15-2026-09-30-live.xlsx"' in disposition


def test_what_we_hold_is_tinted_like_the_cash_bridge(db_session, books):
    """Cash and bank are what a partner opens the pack to find.

    They sat in a plain table while the cash bridge directly above them was
    tinted, so the two figures people actually look for read like every
    movement line on the sheet. Money held is blue, money owed amber.
    """
    _sale(db_session, books, date(2026, 6, 10), 100_000)
    wb, _ = _pack(db_session, books)
    summary = wb["Summary"]

    def row_for(label: str) -> int:
        for r in range(1, summary.max_row + 1):
            if summary.cell(row=r, column=1).value == label:
                return r
        raise AssertionError(f"{label!r} not on the Summary sheet")

    for label in ("Cash in hand", "Bank"):
        cell = summary.cell(row=row_for(label), column=1)
        assert cell.font.bold, f"{label} should be bold"
        assert cell.fill.fgColor.rgb not in (None, "00000000"), (
            f"{label} should be tinted, not left on the default fill"
        )

    # Held and owed must be visually distinct, or the colour says nothing.
    held = summary.cell(row=row_for("Cash in hand"), column=1).fill.fgColor.rgb
    owed = summary.cell(row=row_for("Owed to suppliers"), column=1).fill.fgColor.rgb
    assert held != owed, "money held and money owed should not share a tint"


def test_sales_sheet_carries_a_running_net(db_session, books):
    """A day's net alone is misleading; the carried figure is the answer.

    Expenses arrive in lumps — one supplier invoice can put a day deep into
    the red while the month is well ahead — so the sheet showed a column of
    wild swings with nothing saying where the period actually stood.
    """
    _sale(db_session, books, date(2026, 6, 10), 100_000)
    _sale(db_session, books, date(2026, 6, 20), 250_000)
    wb, _ = _pack(db_session, books)
    sales = wb["Sales"]

    assert sales.cell(row=4, column=5).value == money_header("Running net")

    nets: list[float] = []
    runnings: list[float] = []
    for r in range(5, sales.max_row + 1):
        if sales.cell(row=r, column=1).value in (None, "Total for the period"):
            break
        nets.append(sales.cell(row=r, column=4).value or 0)
        runnings.append(sales.cell(row=r, column=5).value or 0)

    assert nets, "expected daily rows"
    # Each running figure is every net up to and including that day.
    for i, running in enumerate(runnings):
        assert round(running, 2) == round(sum(nets[: i + 1]), 2), f"row {i}"

    # And the period total agrees with the last running figure.
    total_row = next(
        r
        for r in range(5, sales.max_row + 1)
        if sales.cell(row=r, column=1).value == "Total for the period"
    )
    assert round(sales.cell(row=total_row, column=4).value, 2) == round(
        runnings[-1], 2
    )
    assert round(sales.cell(row=total_row, column=2).value, 2) == round(
        sum(
            sales.cell(row=r, column=2).value or 0
            for r in range(5, total_row)
        ),
        2,
    )


def _assert_finished_metric_sheet(ws) -> int:
    """Shared finish: Metric / Amount (₺) header, freeze, accounting money."""
    header_row = None
    expected = ["Metric", money_header()]
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 3)]
        if vals == expected:
            header_row = r
            break
    assert header_row is not None, "Metric / Amount header not found"
    assert ws.freeze_panes == f"A{header_row + 1}"
    assert ws.page_setup.fitToWidth == 1
    money_seen = False
    for r in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=r, column=2)
        if isinstance(cell.value, (int, float)):
            money_seen = True
            assert cell.number_format == MONEY_FORMAT_ACCOUNTING
            assert not isinstance(cell.value, str)
    assert money_seen
    return header_row


def test_summary_sheet_uses_shared_table_finish(db_session, books):
    _sale(db_session, books, date(2026, 6, 10), 100_000)
    wb, _ = _pack(db_session, books)
    _assert_finished_metric_sheet(wb["Summary"])


def test_card_clearing_sheet_uses_shared_table_finish(db_session, books):
    _sale(db_session, books, date(2026, 6, 10), 100_000)
    wb, _ = _pack(db_session, books)
    assert "Card clearing" in wb.sheetnames
    _assert_finished_metric_sheet(wb["Card clearing"])
