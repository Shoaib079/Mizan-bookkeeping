"""Human-readable labels in Excel exports — partner books language only."""

from __future__ import annotations

from app.core.excel.labels import (
    assert_partner_journal_labels_complete,
    format_journal_source,
)
from app.core.ledger.models import JournalEntrySource


def test_journal_source_labels_are_partner_readable() -> None:
    assert format_journal_source(JournalEntrySource.EXPENSE_ENTRY) == "Miscellaneous expense"
    assert format_journal_source(JournalEntrySource.RULE_AUTO) == "Bank transaction"
    assert format_journal_source(JournalEntrySource.SYSTEM) == "Other income"
    assert format_journal_source(JournalEntrySource.MANUAL) == "Adjustment"
    assert format_journal_source(JournalEntrySource.CARD_SALES) == "Card sales"
    assert (
        format_journal_source(JournalEntrySource.POS_COMMISSION_STATEMENT)
        == "Card commission"
    )
    assert (
        format_journal_source(JournalEntrySource.POS_COMMISSION_SWEEP)
        == "Card commission"
    )


def test_every_journal_source_has_a_clear_partner_label() -> None:
    assert_partner_journal_labels_complete()
    for source in JournalEntrySource:
        label = format_journal_source(source)
        assert label
        assert "_" not in label
        lowered = label.lower()
        for banned in ("auto", "rule", "system", "sweep", "batch"):
            assert banned not in lowered.split(), (source.value, label)


def test_fit_columns_expands_for_long_account_names() -> None:
    from datetime import date

    from openpyxl import Workbook

    from app.core.excel.workbook import fit_columns_from_content, write_header_row, write_money

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    header_row = write_header_row(ws, 1, ["Date", "Account", "Amount (₺)"])
    long_account = "5310 — Kart Komisyonu ve Banka Masrafları"
    ws.cell(row=header_row, column=1, value=date(2026, 7, 1))
    ws.cell(row=header_row, column=2, value=long_account)
    write_money(ws, header_row, 3, 12_345)

    fit_columns_from_content(
        ws,
        first_row=1,
        last_row=header_row,
        last_col=3,
        min_widths={1: 12, 2: 20, 3: 14},
        max_widths={1: 14, 2: 52, 3: 16},
    )

    assert ws.column_dimensions["B"].width >= len(long_account)
