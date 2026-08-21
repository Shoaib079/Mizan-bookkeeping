"""S11 — cash flow / KDV / delivery-sales / period-comparison money_cols.

Assert by loading workbooks (openpyxl cell properties), never by grepping
builders. Mutation: wrong header_row → red; drop money_cols → red.
"""

from __future__ import annotations

import uuid
from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from app.core.excel.workbook import MONEY_FORMAT_ACCOUNTING, money_header
from app.features.reports.excel_export import (
    build_cash_flow_xlsx,
    build_delivery_sales_xlsx,
    build_kdv_input_xlsx,
    build_period_comparison_xlsx,
)
from app.features.reports.schema import (
    CashFlowCategoryRead,
    CashFlowRead,
    CashFlowSourceRow,
    DeliverySalesPlatformRow,
    DeliverySalesReportRead,
    KdvInputRateRow,
    KdvInputReportRead,
    PeriodComparisonRead,
    PeriodMetricComparison,
)


def _sheet(data: bytes):
    return load_workbook(BytesIO(data)).active


def _find_header(ws, expected: list[str]) -> int:
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, len(expected) + 1)]
        if vals == expected:
            return r
    raise AssertionError(f"header {expected!r} not found")


def _assert_money_cols(ws, *, header_row: int, money_cols: tuple[int, ...]) -> None:
    assert ws.freeze_panes == f"A{header_row + 1}"
    assert ws.auto_filter.ref is not None
    assert ws.auto_filter.ref.startswith(f"A{header_row}:")
    assert ws.page_setup.fitToWidth == 1
    money_seen = False
    for r in range(header_row + 1, ws.max_row + 1):
        for col in money_cols:
            cell = ws.cell(row=r, column=col)
            if isinstance(cell.value, (int, float)):
                money_seen = True
                assert cell.number_format == MONEY_FORMAT_ACCOUNTING
                assert not isinstance(cell.value, str)
    assert money_seen


def test_cash_flow_xlsx_passes_money_cols() -> None:
    zero = CashFlowCategoryRead(inflows_kurus=0, outflows_kurus=0, net_kurus=0)
    data = build_cash_flow_xlsx(
        CashFlowRead(
            entity_id=uuid.uuid4(),
            from_date=date(2026, 1, 1),
            to_date=date(2026, 1, 31),
            opening_cash_kurus=100_000,
            closing_cash_kurus=80_000,
            net_change_kurus=-20_000,
            operating=CashFlowCategoryRead(
                inflows_kurus=50_000, outflows_kurus=70_000, net_kurus=-20_000
            ),
            investing=zero,
            financing=zero,
            by_source=[
                CashFlowSourceRow(
                    source="pos_daily_summary",
                    category="operating",
                    net_cash_kurus=-20_000,
                )
            ],
            reconciled_to_categories=True,
        )
    )
    ws = _sheet(data)
    header_row = _find_header(ws, ["Metric", money_header()])
    _assert_money_cols(ws, header_row=header_row, money_cols=(2, 3))


def test_kdv_input_xlsx_passes_money_cols() -> None:
    data = build_kdv_input_xlsx(
        KdvInputReportRead(
            entity_id=uuid.uuid4(),
            from_date=date(2026, 5, 1),
            to_date=date(2026, 5, 31),
            rates=[
                KdvInputRateRow(
                    rate_percent=20.0,
                    base_kurus=200_000,
                    vat_kurus=40_000,
                    invoice_count=1,
                )
            ],
            total_base_kurus=200_000,
            total_vat_kurus=40_000,
            invoice_count=1,
        )
    )
    ws = _sheet(data)
    header_row = _find_header(
        ws, ["Rate (%)", money_header("Base"), money_header("VAT"), "Invoice count"]
    )
    _assert_money_cols(ws, header_row=header_row, money_cols=(2, 3))


def test_delivery_sales_xlsx_passes_money_cols() -> None:
    data = build_delivery_sales_xlsx(
        DeliverySalesReportRead(
            entity_id=uuid.uuid4(),
            from_date=date(2026, 11, 1),
            to_date=date(2026, 11, 30),
            platforms=[
                DeliverySalesPlatformRow(
                    delivery_platform_id=uuid.uuid4(),
                    platform_name="Getir",
                    is_active=True,
                    gross_kurus=180_000,
                    report_count=1,
                )
            ],
            total_gross_kurus=180_000,
        )
    )
    ws = _sheet(data)
    header_row = _find_header(
        ws, ["Platform", "Active", money_header("Gross"), "Report count"]
    )
    _assert_money_cols(ws, header_row=header_row, money_cols=(3,))


def test_period_comparison_xlsx_passes_money_cols() -> None:
    data = build_period_comparison_xlsx(
        PeriodComparisonRead(
            entity_id=uuid.uuid4(),
            current_from=date(2026, 1, 1),
            current_to=date(2026, 1, 31),
            prior_from=date(2025, 12, 1),
            prior_to=date(2025, 12, 31),
            metrics=[
                PeriodMetricComparison(
                    key="net_income",
                    label="Net income",
                    current_kurus=100_000,
                    prior_kurus=80_000,
                    change_kurus=20_000,
                    change_percent=25.0,
                )
            ],
        )
    )
    ws = _sheet(data)
    header_row = _find_header(
        ws,
        [
            "Metric",
            money_header("Current"),
            money_header("Prior"),
            money_header("Change"),
            "Change (%)",
        ],
    )
    _assert_money_cols(ws, header_row=header_row, money_cols=(2, 3, 4))
