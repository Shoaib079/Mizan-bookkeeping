"""Group-sale credit rows show menu · pax × rate on screen, PDF, and Excel."""

from __future__ import annotations

import uuid
from datetime import date
from io import BytesIO
from types import SimpleNamespace

import fitz
import pytest
from openpyxl import load_workbook
from sqlalchemy import func, select

from app.core.chart_of_accounts.seed import seed_default_chart
from app.core.receivables.models import CustomerLedgerEntry
from app.core.receivables.types import CustomerMovementType
from app.db.session import entity_context
from app.features.customers import ledger_export, service as customer_service
from app.features.customers.models import Customer
from app.features.entities import service as entity_service
from app.features.group_sales.ledger_display_description import (
    build_group_sale_ledger_display_description,
)
from app.features.group_sales.models import GroupMenu, GroupSaleLine
from app.features.group_sales.schema import GroupSaleCreate, GroupSaleLineInput
from app.features.group_sales import service as group_sales_service


ACTOR_ID = uuid.UUID("00000000-0000-4000-8000-000000000001")
HELPER_SRC = (
    __import__("pathlib").Path(__file__).resolve().parents[1]
    / "app/features/group_sales/ledger_display_description.py"
)


@pytest.fixture
def rich_desc_setup(db_session, restaurant_a):
    seed_default_chart(db_session, restaurant_a.id)
    with entity_context(db_session, restaurant_a.id):
        customer = Customer(name="Rich Desc Agency")
        db_session.add(customer)
        veg = GroupMenu(name="Veg Menu 1")
        metro = GroupMenu(name="Metro Menu")
        db_session.add_all([customer, veg, metro])
        db_session.commit()
        db_session.refresh(customer)
        db_session.refresh(veg)
        db_session.refresh(metro)
    return {
        "entity_id": restaurant_a.id,
        "customer_id": customer.id,
        "veg_menu_id": veg.id,
        "metro_menu_id": metro.id,
    }


def _credit_sale_descriptions(ledger) -> list[str]:
    return [
        entry.description
        for entry in ledger.entries
        if entry.movement_type == CustomerMovementType.CREDIT_SALE
        and entry.display_kind.value == "effective"
    ]


def _excel_descriptions(data: bytes) -> list[str]:
    wb = load_workbook(BytesIO(data), data_only=True)
    ws = wb.active
    header_row = desc_col = None
    for r in range(1, 30):
        for c in range(1, 10):
            val = ws.cell(row=r, column=c).value
            if val == "Description":
                header_row, desc_col = r, c
                break
        if header_row is not None:
            break
    assert header_row is not None and desc_col is not None
    out: list[str] = []
    for r in range(header_row + 1, ws.max_row + 1):
        movement = ws.cell(row=r, column=2).value
        if movement and "Credit" in str(movement):
            cell = ws.cell(row=r, column=desc_col).value
            if cell:
                out.append(str(cell))
    return out


def _pdf_text(data: bytes) -> str:
    with fitz.open(stream=data, filetype="pdf") as doc:
        return "\n".join(page.get_text() for page in doc)


def _export_surfaces(db_session, setup, ledger, entity_name, customer_name):
    xlsx = ledger_export.build_customer_ledger_xlsx(
        entity_name=entity_name,
        customer_name=customer_name,
        ledger=ledger,
    )
    pdf = ledger_export.build_customer_ledger_pdf(
        entity_name=entity_name,
        customer_name=customer_name,
        ledger=ledger,
    )
    return _credit_sale_descriptions(ledger), _excel_descriptions(xlsx), _pdf_text(pdf)


def test_single_line_rich_description_api_pdf_excel(db_session, rich_desc_setup):
    entity_id = rich_desc_setup["entity_id"]
    customer_id = rich_desc_setup["customer_id"]
    expected = "Veg Menu 1 · 10 pax × $12.00"

    sale = group_sales_service.post_group_sale(
        db_session,
        entity_id,
        GroupSaleCreate(
            customer_id=customer_id,
            sale_date=date(2026, 8, 1),
            description="Group sale",
            currency="USD",
            fx_rate_used=3_500,
            actor_id=ACTOR_ID,
            lines=[
                GroupSaleLineInput(
                    group_menu_id=rich_desc_setup["veg_menu_id"],
                    pax=10,
                    rate_per_person_minor=1_200,
                ),
            ],
        ),
    )

    with entity_context(db_session, entity_id):
        stored = db_session.get(CustomerLedgerEntry, sale.customer_ledger_entry_id)
        assert stored is not None
        assert stored.description == "Group sale"

    ledger = customer_service.get_customer_ledger(db_session, entity_id, customer_id)
    entity = entity_service.get_entity(db_session, entity_id)
    assert entity is not None
    api, excel, pdf = _export_surfaces(
        db_session, rich_desc_setup, ledger, entity.name, "Rich Desc Agency"
    )
    assert api == [expected]
    assert excel == [expected]
    assert expected in pdf


def test_multi_line_rich_description_joined(db_session, rich_desc_setup):
    entity_id = rich_desc_setup["entity_id"]
    customer_id = rich_desc_setup["customer_id"]
    expected = (
        "Metro Menu · 8 pax × $15.00 + Veg Menu 1 · 10 pax × $12.00"
    )

    group_sales_service.post_group_sale(
        db_session,
        entity_id,
        GroupSaleCreate(
            customer_id=customer_id,
            sale_date=date(2026, 8, 2),
            description="Group sale",
            currency="USD",
            fx_rate_used=3_500,
            actor_id=ACTOR_ID,
            lines=[
                GroupSaleLineInput(
                    group_menu_id=rich_desc_setup["veg_menu_id"],
                    pax=10,
                    rate_per_person_minor=1_200,
                ),
                GroupSaleLineInput(
                    group_menu_id=rich_desc_setup["metro_menu_id"],
                    pax=8,
                    rate_per_person_minor=1_500,
                ),
            ],
        ),
    )

    ledger = customer_service.get_customer_ledger(db_session, entity_id, customer_id)
    entity = entity_service.get_entity(db_session, entity_id)
    assert entity is not None
    api, excel, pdf = _export_surfaces(
        db_session, rich_desc_setup, ledger, entity.name, "Rich Desc Agency"
    )
    assert api == [expected]
    assert excel == [expected]
    assert expected in pdf


def test_rich_description_appends_owner_note(db_session, rich_desc_setup):
    entity_id = rich_desc_setup["entity_id"]
    customer_id = rich_desc_setup["customer_id"]
    note = "May tour group"
    expected = f"Veg Menu 1 · 10 pax × $12.00 — {note}"

    group_sales_service.post_group_sale(
        db_session,
        entity_id,
        GroupSaleCreate(
            customer_id=customer_id,
            sale_date=date(2026, 5, 10),
            description=note,
            currency="USD",
            fx_rate_used=3_500,
            actor_id=ACTOR_ID,
            lines=[
                GroupSaleLineInput(
                    group_menu_id=rich_desc_setup["veg_menu_id"],
                    pax=10,
                    rate_per_person_minor=1_200,
                ),
            ],
        ),
    )

    ledger = customer_service.get_customer_ledger(db_session, entity_id, customer_id)
    entity = entity_service.get_entity(db_session, entity_id)
    assert entity is not None
    api, excel, pdf = _export_surfaces(
        db_session, rich_desc_setup, ledger, entity.name, "Rich Desc Agency"
    )
    assert api == [expected]
    assert excel == [expected]
    assert expected in pdf


def test_rich_description_whitespace_note_treated_as_default():
    sale = SimpleNamespace(description="   ", currency="USD")
    line = SimpleNamespace(
        id=1,
        menu_name_snapshot="Veg Menu 1",
        pax=10,
        rate_per_person_minor=1_200,
    )
    assert (
        build_group_sale_ledger_display_description(sale, [line])
        == "Veg Menu 1 · 10 pax × $12.00"
    )


def test_rich_description_deposit_paid_note(db_session, rich_desc_setup):
    entity_id = rich_desc_setup["entity_id"]
    customer_id = rich_desc_setup["customer_id"]
    note = "deposit paid"
    expected = f"Veg Menu 1 · 10 pax × $12.00 — {note}"

    group_sales_service.post_group_sale(
        db_session,
        entity_id,
        GroupSaleCreate(
            customer_id=customer_id,
            sale_date=date(2026, 8, 3),
            description=note,
            currency="USD",
            fx_rate_used=3_500,
            actor_id=ACTOR_ID,
            lines=[
                GroupSaleLineInput(
                    group_menu_id=rich_desc_setup["veg_menu_id"],
                    pax=10,
                    rate_per_person_minor=1_200,
                ),
            ],
        ),
    )

    ledger = customer_service.get_customer_ledger(db_session, entity_id, customer_id)
    entity = entity_service.get_entity(db_session, entity_id)
    assert entity is not None
    api, excel, pdf = _export_surfaces(
        db_session, rich_desc_setup, ledger, entity.name, "Rich Desc Agency"
    )
    assert api == [expected]
    assert excel == [expected]
    assert expected in pdf


def test_rich_description_no_lines_fallback():
    sale = SimpleNamespace(description="Group sale", currency="USD")
    assert build_group_sale_ledger_display_description(sale, []) == "Group sale"


def test_historical_posted_sale_rich_without_migration(db_session, rich_desc_setup):
    """Sale posted with flat stored description; lines already in DB → rich label."""
    entity_id = rich_desc_setup["entity_id"]
    customer_id = rich_desc_setup["customer_id"]
    expected = "Veg Menu 1 · 10 pax × $12.00"

    sale = group_sales_service.post_group_sale(
        db_session,
        entity_id,
        GroupSaleCreate(
            customer_id=customer_id,
            sale_date=date(2026, 5, 15),
            description="Group sale",
            currency="USD",
            fx_rate_used=3_500,
            actor_id=ACTOR_ID,
            lines=[
                GroupSaleLineInput(
                    group_menu_id=rich_desc_setup["veg_menu_id"],
                    pax=10,
                    rate_per_person_minor=1_200,
                ),
            ],
        ),
    )

    with entity_context(db_session, entity_id):
        stored = db_session.get(CustomerLedgerEntry, sale.customer_ledger_entry_id)
        assert stored is not None
        assert stored.description == "Group sale"
        line_count = db_session.scalar(
            select(func.count())
            .select_from(GroupSaleLine)
            .where(GroupSaleLine.group_sale_id == sale.id)
        )
        assert line_count == 1

    ledger = customer_service.get_customer_ledger(db_session, entity_id, customer_id)
    assert _credit_sale_descriptions(ledger) == [expected]


def test_mutation_flat_group_sale_for_lined_sale():
    src = HELPER_SRC.read_text()
    join_line = 'body = " + ".join(_format_line(line, group_sale.currency) for line in ordered)'
    assert join_line in src
    broken = src.replace(join_line, 'body = "Group sale"')
    assert join_line not in broken
    namespace: dict = {}
    exec(compile(broken, str(HELPER_SRC), "exec"), namespace)
    build = namespace["build_group_sale_ledger_display_description"]
    line = SimpleNamespace(
        id=1,
        menu_name_snapshot="Veg Menu 1",
        pax=10,
        rate_per_person_minor=1_200,
    )
    sale = SimpleNamespace(description="Group sale", currency="USD")
    assert build(sale, [line]) == "Group sale"
