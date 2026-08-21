"""S15 Part B — customer ledger Running on Excel/PDF matches the screen source.

Running used to be hard-coded None in the export builder while the partner
export passed entry.running_balance_kurus. Assert by loading the workbook and
rendering the PDF — never by grepping the builder.
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

import fitz
import pytest
from openpyxl import load_workbook

from app.core.auth.types import EntityRole
from app.core.chart_of_accounts.seed import seed_default_chart
from app.core.money import format_try
from app.features.auth import service as auth_service
from app.features.auth.schema import MembershipCreate, UserCreate
from app.features.banking import service as banking_service
from app.features.banking.models import MoneyAccountKind
from app.features.banking.schema import MoneyAccountCreate
from app.features.customers import ledger_export, service as customer_service
from app.features.customers.schema import (
    CreditSaleCreate,
    CustomerCreate,
    CustomerPaymentCreate,
)
from app.features.entities import service as entity_service


JUNE = date(2026, 6, 10)
SALE_1 = 100_000
SALE_2 = 50_000
PAYMENT = 80_000


@pytest.fixture
def books(db_session, restaurant_a):
    seed_default_chart(db_session, restaurant_a.id)
    owner = auth_service.create_user(
        db_session,
        UserCreate(email="s15b-customer@example.com", display_name="Owner"),
    )
    auth_service.add_entity_member(
        db_session,
        restaurant_a.id,
        MembershipCreate(user_id=owner.id, role=EntityRole.OWNER),
    )
    cash = banking_service.create_money_account(
        db_session,
        restaurant_a.id,
        MoneyAccountCreate(account_kind=MoneyAccountKind.CASH, name="Till"),
    )
    return {
        "entity_id": restaurant_a.id,
        "owner_id": owner.id,
        "cash_gl": cash.gl_account_id,
    }


def _customer_with_credit_sales_and_payment(db_session, books):
    """Group/agency credit sales + a payment — the rows that used to export —."""
    entity_id = books["entity_id"]
    customer = customer_service.create_customer(
        db_session,
        entity_id,
        CustomerCreate(name="Agency Tour"),
    )
    customer_id = customer.id
    customer_name = customer.name
    # Credit sales with pax/rate — same customer-ledger movement type a group
    # sale posts; Running must accumulate on these, not only on payments.
    for amount, label, pax in (
        (SALE_1, "Trip A", 10),
        (SALE_2, "Trip B", 5),
    ):
        customer_service.record_credit_sale(
            db_session,
            entity_id,
            customer_id,
            CreditSaleCreate(
                sale_date=JUNE,
                amount_kurus=amount,
                description=label,
                actor_id=books["owner_id"],
                pax=pax,
                rate_per_person_kurus=amount // pax,
            ),
        )
    customer_service.record_customer_payment(
        db_session,
        entity_id,
        customer_id,
        CustomerPaymentCreate(
            payment_date=JUNE,
            amount_kurus=PAYMENT,
            description="Partial receipt",
            payment_account_id=books["cash_gl"],
            actor_id=books["owner_id"],
        ),
    )
    ledger = customer_service.get_customer_ledger(
        db_session, entity_id, customer_id
    )
    entity = entity_service.get_entity(db_session, entity_id)
    assert entity is not None
    return customer_id, customer_name, ledger, entity.name


def test_customer_ledger_screen_running_accumulates(db_session, books):
    _, _, ledger, _ = _customer_with_credit_sales_and_payment(db_session, books)
    effective = [
        e for e in ledger.entries if e.display_kind.value == "effective"
    ]
    assert len(effective) == 3
    runnings = [e.running_balance_kurus for e in effective]
    assert runnings == [SALE_1, SALE_1 + SALE_2, SALE_1 + SALE_2 - PAYMENT]
    assert runnings[-1] == ledger.balance_kurus


def test_customer_ledger_excel_running_matches_screen(db_session, books):
    _, customer_name, ledger, entity_name = _customer_with_credit_sales_and_payment(
        db_session, books
    )
    data = ledger_export.build_customer_ledger_xlsx(
        entity_name=entity_name,
        customer_name=customer_name,
        ledger=ledger,
    )
    wb = load_workbook(BytesIO(data), data_only=True)
    ws = wb.active
    header_row = None
    running_col = None
    for r in range(1, 25):
        for c in range(1, 8):
            val = ws.cell(row=r, column=c).value
            if val and "Running" in str(val):
                header_row = r
                running_col = c
                break
        if header_row is not None:
            break
    assert header_row is not None and running_col is not None

    screen_runnings = [
        e.running_balance_kurus
        for e in ledger.entries
        if e.display_kind.value == "effective"
    ]
    export_runnings = []
    for r in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=r, column=running_col).value
        if isinstance(cell, (int, float)):
            export_runnings.append(int(round(cell * 100)))

    assert export_runnings == screen_runnings
    assert export_runnings[-1] == ledger.balance_kurus


def test_customer_ledger_pdf_running_matches_screen(db_session, books):
    _, customer_name, ledger, entity_name = _customer_with_credit_sales_and_payment(
        db_session, books
    )
    data = ledger_export.build_customer_ledger_pdf(
        entity_name=entity_name,
        customer_name=customer_name,
        ledger=ledger,
    )
    with fitz.open(stream=data, filetype="pdf") as doc:
        text = "\n".join(page.get_text() for page in doc)
    assert "Running" in text
    assert format_try(ledger.balance_kurus) in text
    assert format_try(SALE_1) in text
    # Em-dash placeholder must not be the only Running cell content for sales.
    # (PDF uses format_try for amounts; blank running was "—".)
    assert text.count("—") < 3 or format_try(SALE_1 + SALE_2) in text
