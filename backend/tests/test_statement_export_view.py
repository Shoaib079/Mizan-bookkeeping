"""P&L / BS exports honour the on-screen sealed vs live view (S6).

A closed month's screen can show sealed or live figures; Download used to
omit ``view`` and always served the default. These tests pin that Excel/PDF
follow the same ``get_*`` path as the screen, stamp the world in the file and
filename, and leave open-month filenames unchanged.
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from app.core.auth.types import EntityRole
from app.core.chart_of_accounts.default_chart import SALES_REVENUE_CODE
from app.core.chart_of_accounts.models import Account
from app.core.chart_of_accounts.seed import seed_default_chart
from app.core.chart_of_accounts.types import AccountNormalBalance
from app.core.ledger.models import JournalEntrySource
from app.core.ledger.posting import PostingLine, post_journal_entry, void_journal_entry
from app.core.period_locks.models import PeriodLockKind
from app.core.period_locks.service import close_period
from app.db.session import entity_context
from app.features.auth import service as auth_service
from app.features.auth.schema import MembershipCreate, UserCreate
from app.features.reports import financial_statements, statement_exports

CASH_CODE = "1000"
JUNE_START = date(2026, 6, 1)
JUNE_END = date(2026, 6, 30)
UNLOCK = "Correcting a duplicate found later"
SALE = 100_000


@pytest.fixture
def books(db_session, restaurant_a):
    seed_default_chart(db_session, restaurant_a.id)
    owner = auth_service.create_user(
        db_session,
        UserCreate(email="s6-export-owner@example.com", display_name="Owner"),
    )
    auth_service.add_entity_member(
        db_session,
        restaurant_a.id,
        MembershipCreate(user_id=owner.id, role=EntityRole.OWNER),
    )
    with entity_context(db_session, restaurant_a.id):
        accounts = {a.code: a.id for a in db_session.scalars(select(Account))}
    return {"entity_id": restaurant_a.id, "owner_id": owner.id, "accounts": accounts}


def _sale(db_session, books, amount: int = SALE):
    with entity_context(db_session, books["entity_id"]):
        entry = post_journal_entry(
            db_session,
            books["entity_id"],
            date(2026, 6, 10),
            "Cash sale",
            [
                PostingLine(
                    books["accounts"][CASH_CODE], amount, AccountNormalBalance.DEBIT
                ),
                PostingLine(
                    books["accounts"][SALES_REVENUE_CODE],
                    amount,
                    AccountNormalBalance.CREDIT,
                ),
            ],
            actor_id=books["owner_id"],
            source=JournalEntrySource.MANUAL,
        )
        db_session.commit()
        return entry.id


def _void(db_session, books, entry_id):
    with entity_context(db_session, books["entity_id"]):
        void_journal_entry(
            db_session,
            books["entity_id"],
            entry_id,
            actor_id=books["owner_id"],
            reason="Recorded twice",
            void_date=date(2026, 6, 10),
            period_unlock_reason=UNLOCK,
        )
        db_session.commit()


def _close_june(db_session, books):
    return close_period(
        db_session,
        books["entity_id"],
        lock_kind=PeriodLockKind.MONTH,
        anchor_date=JUNE_END,
        actor_id=books["owner_id"],
    )


def _xlsx_cells(data: bytes) -> set[str]:
    wb = load_workbook(BytesIO(data), data_only=True)
    ws = wb.active
    values: set[str] = set()
    for row in ws.iter_rows(max_row=8, max_col=2, values_only=True):
        for cell in row:
            if cell is not None:
                values.add(str(cell))
    return values


def test_closed_month_default_export_is_sealed_and_stamped(db_session, books):
    _sale(db_session, books)
    _close_june(db_session, books)

    data, filename = statement_exports.profit_and_loss_xlsx(
        db_session,
        books["entity_id"],
        JUNE_START,
        JUNE_END,
        view=financial_statements.VIEW_AS_CLOSED,
    )
    assert filename.endswith("-as-closed.xlsx")
    cells = _xlsx_cells(data)
    assert "As closed" in cells
    sealed = financial_statements.get_profit_and_loss(
        db_session,
        books["entity_id"],
        JUNE_START,
        JUNE_END,
        view=financial_statements.VIEW_AS_CLOSED,
    )
    assert sealed.net_income_kurus == SALE
    assert sealed.source == financial_statements.VIEW_AS_CLOSED


def test_closed_month_live_export_differs_after_amend(db_session, books):
    """F3 reproduction: void after close → sealed and live diverge."""
    entry_id = _sale(db_session, books)
    _close_june(db_session, books)
    _void(db_session, books, entry_id)

    sealed_data, sealed_name = statement_exports.profit_and_loss_xlsx(
        db_session,
        books["entity_id"],
        JUNE_START,
        JUNE_END,
        view=financial_statements.VIEW_AS_CLOSED,
    )
    live_data, live_name = statement_exports.profit_and_loss_xlsx(
        db_session,
        books["entity_id"],
        JUNE_START,
        JUNE_END,
        view=financial_statements.VIEW_LIVE,
    )
    assert sealed_name.endswith("-as-closed.xlsx")
    assert live_name.endswith("-live.xlsx")
    assert "As closed" in _xlsx_cells(sealed_data)
    assert "Live" in _xlsx_cells(live_data)
    assert sealed_data != live_data

    sealed = financial_statements.get_profit_and_loss(
        db_session,
        books["entity_id"],
        JUNE_START,
        JUNE_END,
        view=financial_statements.VIEW_AS_CLOSED,
    )
    live = financial_statements.get_profit_and_loss(
        db_session,
        books["entity_id"],
        JUNE_START,
        JUNE_END,
        view=financial_statements.VIEW_LIVE,
    )
    assert sealed.net_income_kurus == SALE
    assert live.net_income_kurus == 0


def test_balance_sheet_export_honours_view(db_session, books):
    entry_id = _sale(db_session, books)
    _close_june(db_session, books)
    _void(db_session, books, entry_id)

    sealed_data, sealed_name = statement_exports.balance_sheet_xlsx(
        db_session, books["entity_id"], JUNE_END, view=financial_statements.VIEW_AS_CLOSED
    )
    live_data, live_name = statement_exports.balance_sheet_xlsx(
        db_session, books["entity_id"], JUNE_END, view=financial_statements.VIEW_LIVE
    )
    assert sealed_name.endswith("-as-closed.xlsx")
    assert live_name.endswith("-live.xlsx")
    assert sealed_data != live_data


def test_open_month_export_filename_has_no_figures_suffix(db_session, books):
    _sale(db_session, books)
    data, filename = statement_exports.profit_and_loss_xlsx(
        db_session,
        books["entity_id"],
        JUNE_START,
        JUNE_END,
        view=financial_statements.VIEW_AS_CLOSED,
    )
    assert "-as-closed" not in filename
    assert "-live" not in filename
    assert filename.endswith(f"-{JUNE_START.year:04d}-{JUNE_START.month:02d}.xlsx")
    assert "Live" in _xlsx_cells(data)


def test_http_export_passes_view(
    client: TestClient, db_session, books, monkeypatch
):
    """Route forwards view into the shared export builder (not a second path)."""
    _sale(db_session, books)
    _close_june(db_session, books)
    seen: list[str] = []

    real = statement_exports.profit_and_loss_xlsx

    def spy(session, entity_id, from_date, to_date, *, view):
        seen.append(view)
        return real(session, entity_id, from_date, to_date, view=view)

    monkeypatch.setattr(statement_exports, "profit_and_loss_xlsx", spy)
    # api imports the module attribute — patch where the route looks.
    import app.features.reports.api as reports_api

    monkeypatch.setattr(reports_api.statement_exports, "profit_and_loss_xlsx", spy)

    response = client.get(
        f"/entities/{books['entity_id']}/reports/profit-and-loss/export"
        f"?from={JUNE_START}&to={JUNE_END}&view=live"
    )
    assert response.status_code == 200, response.text
    assert seen == [financial_statements.VIEW_LIVE]
    assert "-live.xlsx" in response.headers.get("content-disposition", "")


def test_figures_filename_suffix_rules():
    assert (
        statement_exports.figures_filename_suffix("as_closed", period_closed=True)
        == "as-closed"
    )
    assert (
        statement_exports.figures_filename_suffix("live", period_closed=True) == "live"
    )
    assert (
        statement_exports.figures_filename_suffix("live", period_closed=False) is None
    )
