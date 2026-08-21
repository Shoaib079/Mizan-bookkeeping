"""S8 — subledger / activity / delivery / POS Excel use shared finish_data_table.

Assert by loading the workbook (openpyxl cell properties), never by grepping
builders. Mutation: wrong header_row → red; drop money_cols → red.
"""

from __future__ import annotations

import uuid
from datetime import date
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.core.excel.workbook import MONEY_FORMAT_ACCOUNTING, money_header
from app.features.delivery.excel_export import build_delivery_activity_xlsx
from app.features.payables.activity_excel import build_supplier_activity_xlsx
from app.features.payables.schema import SupplierActivityRead, SupplierActivityRow
from app.features.pos.excel_export import build_pos_daily_summaries_xlsx
from app.features.reports.subledger_export import (
    SubledgerExport,
    SubledgerRow,
    build_subledger_xlsx,
)


def _sheet(data: bytes, name: str | None = None):
    wb = load_workbook(BytesIO(data))
    return wb[name] if name else wb.active


def _find_header(ws, expected: list[str]) -> int:
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, len(expected) + 1)]
        if vals == expected:
            return r
    raise AssertionError(f"header {expected!r} not found")


def _assert_money_cols(ws, *, header_row: int, money_cols: tuple[int, ...]) -> None:
    assert ws.freeze_panes == f"A{header_row + 1}"
    assert ws.auto_filter.ref is not None
    assert ws.auto_filter.ref.startswith(f"A{header_row}:")
    assert ws.page_setup.fitToWidth == 1
    money_seen = False
    for r in range(header_row + 1, ws.max_row + 1):
        for col in money_cols:
            cell = ws.cell(row=r, column=col)
            if isinstance(cell.value, (int, float)):
                money_seen = True
                assert cell.number_format == MONEY_FORMAT_ACCOUNTING
                assert not isinstance(cell.value, str)
    assert money_seen


def test_subledger_xlsx_uses_shared_table_finish() -> None:
    data = build_subledger_xlsx(
        SubledgerExport(
            entity_name="India Gate",
            subject_name="Ali",
            ledger_label="Partner ledger",
            sheet_name="Partner",
            summary=[("Capital", 50_000)],
            rows=[
                SubledgerRow(
                    movement_date=date(2026, 6, 30),
                    movement="Profit allocation",
                    description="June share",
                    amount_minor=100_000,
                    running_minor=150_000,
                    status="Effective",
                )
            ],
        )
    )
    ws = _sheet(data)
    headers = [
        "Date",
        "Movement",
        "Description",
        money_header("Amount"),
        money_header("Running"),
        "Status",
    ]
    header_row = _find_header(ws, headers)
    _assert_money_cols(ws, header_row=header_row, money_cols=(4, 5))


def test_supplier_activity_xlsx_uses_shared_table_finish() -> None:
    sid = uuid.uuid4()
    data = build_supplier_activity_xlsx(
        SupplierActivityRead(
            supplier_id=sid,
            supplier_name="Metro",
            supplier_vkn="1234567890",
            from_date=date(2026, 1, 1),
            to_date=date(2026, 1, 31),
            opening_balance_kurus=0,
            closing_balance_kurus=120_000,
            total_invoices_gross_kurus=120_000,
            total_payments_kurus=0,
            total_vat_kurus=20_000,
            rows=[
                SupplierActivityRow(
                    movement_date=date(2026, 1, 10),
                    movement_kind="invoice",
                    movement_label="Invoice",
                    document_ref="INV-1",
                    detail="Goods",
                    net_kurus=100_000,
                    vat_kurus=20_000,
                    amount_kurus=120_000,
                    balance_kurus=120_000,
                )
            ],
        ),
        entity_name="Restaurant A",
    )
    ws = _sheet(data)
    headers = [
        "Date",
        "Movement",
        "Document / ref",
        "Detail",
        money_header("Net"),
        money_header("VAT"),
        money_header(),
        "Bank",
        "Receipt",
        money_header("Balance"),
    ]
    header_row = _find_header(ws, headers)
    _assert_money_cols(ws, header_row=header_row, money_cols=(5, 6, 7, 10))


def test_pos_daily_summaries_xlsx_uses_shared_table_finish() -> None:
    data = build_pos_daily_summaries_xlsx(
        entity_name="Restaurant A",
        from_date=date(2026, 3, 1),
        to_date=date(2026, 3, 31),
        review_label="pending",
        summaries=[
            SimpleNamespace(
                summary_date=date(2026, 3, 15),
                status="pending_review",
                cash_kurus=10_000,
                card_kurus=-2_500,
                total_kurus=7_500,
                z_report_kurus=7_500,
                review_reason=None,
                posted_at=None,
            )
        ],
    )
    ws = _sheet(data)
    headers = [
        "Date",
        "Status",
        money_header("Cash"),
        money_header("Card"),
        money_header("Total"),
        money_header("Z report"),
        "Review reason",
        "Posted at",
    ]
    header_row = _find_header(ws, headers)
    _assert_money_cols(ws, header_row=header_row, money_cols=(3, 4, 5, 6))


def test_delivery_activity_xlsx_uses_shared_table_finish() -> None:
    data = build_delivery_activity_xlsx(
        entity_name="Restaurant A",
        from_date=date(2026, 11, 1),
        to_date=date(2026, 11, 30),
        platform_label="All",
        sales=[
            SimpleNamespace(
                platform_name="Getir",
                period_start=date(2026, 11, 1),
                period_end=date(2026, 11, 30),
                gross_kurus=180_000,
                status="posted",
                description="November",
            )
        ],
        settlements=[
            SimpleNamespace(
                platform_name="Getir",
                settlement_date=date(2026, 11, 20),
                amount_kurus=150_000,
                description="Payout",
            )
        ],
    )
    sales = _sheet(data, "Sales")
    sales_header = _find_header(
        sales,
        [
            "Platform",
            "Period from",
            "Period to",
            money_header("Gross"),
            "Status",
            "Description",
        ],
    )
    _assert_money_cols(sales, header_row=sales_header, money_cols=(4,))

    settle = _sheet(data, "Settlements")
    settle_header = _find_header(
        settle,
        ["Platform", "Date", money_header("Amount"), "Description"],
    )
    _assert_money_cols(settle, header_row=settle_header, money_cols=(3,))
