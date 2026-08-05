"""Excel export for Phase 7 reports (Phase 7 Slice 7)."""

from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from app.core.chart_of_accounts.seed import seed_default_chart
from app.core.cash.posting import post_cash_movement
from app.core.chart_of_accounts.default_chart import SALES_REVENUE_CODE
from app.core.chart_of_accounts.models import Account
from app.core.pos import posting as pos_posting
from app.db.session import entity_context
from app.features.banking import service as banking_service
from app.features.banking.models import MoneyAccountKind
from app.features.banking.schema import MoneyAccountCreate
from app.features.cash.models import CashMovementDirection
from app.features.reports import excel_export
from app.features.reports import financial_statements
from tests.delivery_helpers import ACTOR_ID
from tests.test_financial_statements import (
    PERIOD_END,
    PERIOD_START,
    _post_period_sales,
    _post_rent_expense,
)
from tests.test_kdv_input_report import (
    _post_supplier_draft,
    _supplier,
    _supplier_draft,
)

XLSX_CONTENT_TYPE = excel_export.XLSX_CONTENT_TYPE


@pytest.fixture
def export_setup(db_session, restaurant_a):
    seed_default_chart(db_session, restaurant_a.id)
    drawer = banking_service.create_money_account(
        db_session,
        restaurant_a.id,
        MoneyAccountCreate(account_kind=MoneyAccountKind.CASH, name="Main Drawer"),
    )
    with entity_context(db_session, restaurant_a.id):
        accounts = {a.code: a.id for a in db_session.scalars(select(Account))}
    return {
        "entity_id": restaurant_a.id,
        "drawer": drawer,
        "accounts": accounts,
    }


@pytest.fixture
def kdv_export_setup(db_session, restaurant_a):
    seed_default_chart(db_session, restaurant_a.id)
    with entity_context(db_session, restaurant_a.id):
        accounts = {a.code: a.id for a in db_session.scalars(select(Account))}
    return {"entity_id": restaurant_a.id, "accounts": accounts}


def _load_sheet(data: bytes):
    wb = load_workbook(BytesIO(data))
    return wb.active


def _cell_values(ws, row: int) -> list:
    return [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]


def _assert_xlsx_export(response, *, header_text: str) -> None:
    assert response.status_code == 200
    assert response.headers["content-type"] == XLSX_CONTENT_TYPE
    ws = _load_sheet(response.content)
    found_header = False
    for row in range(1, ws.max_row + 1):
        if header_text in _cell_values(ws, row):
            found_header = True
            break
    assert found_header, f"Expected header {header_text!r} in worksheet"


def test_profit_and_loss_export(
    db_session, client: TestClient, export_setup
) -> None:
    setup = export_setup
    _post_period_sales(db_session, setup)
    _post_rent_expense(
        db_session, setup, amount_kurus=20_000, expense_date=date(2026, 1, 16)
    )

    response = client.get(
        f"/entities/{setup['entity_id']}/reports/profit-and-loss/export",
        params={"from": "2026-01-01", "to": "2026-01-31"},
    )
    _assert_xlsx_export(response, header_text="Code")

    pl_json = client.get(
        f"/entities/{setup['entity_id']}/reports/profit-and-loss",
        params={"from": "2026-01-01", "to": "2026-01-31"},
    ).json()
    pl_report = financial_statements.get_profit_and_loss(
        db_session, setup["entity_id"], PERIOD_START, PERIOD_END
    )
    ws = _load_sheet(response.content)
    data_rows = 0
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value in {
            acc.code for acc in pl_report.accounts
        }:
            data_rows += 1
    assert data_rows == len(pl_json["accounts"])


def test_profit_and_loss_export_invalid_date_range(
    client: TestClient, export_setup
) -> None:
    response = client.get(
        f"/entities/{export_setup['entity_id']}/reports/profit-and-loss/export",
        params={"from": "2026-02-01", "to": "2026-01-01"},
    )
    assert response.status_code == 422


def test_balance_sheet_export(
    db_session, client: TestClient, export_setup
) -> None:
    setup = export_setup
    _post_period_sales(db_session, setup)

    response = client.get(
        f"/entities/{setup['entity_id']}/reports/balance-sheet/export",
        params={"as_of": "2026-01-31"},
    )
    _assert_xlsx_export(response, header_text="Assets")

    disposition = response.headers.get("content-disposition", "")
    assert 'filename="restaurant-a-balance-sheet-2026-01-31.xlsx"' in disposition


def test_kdv_input_export(
    client: TestClient, db_session, restaurant_a, kdv_export_setup
) -> None:
    setup = kdv_export_setup
    supplier_id = _supplier(db_session, restaurant_a)
    expense_id = setup["accounts"]["5200"]
    draft = _supplier_draft(
        db_session,
        setup["entity_id"],
        supplier_id,
        invoice_date=date(2026, 5, 1),
        invoice_number="XLSX-INV",
        net_kurus=200_000,
        gross_kurus=240_000,
        vat_breakdown=[
            {"rate_percent": 20, "base_kurus": 200_000, "vat_kurus": 40_000},
        ],
        file_fingerprint="kdv-xlsx-fp",
    )
    _post_supplier_draft(db_session, setup["entity_id"], draft.id, expense_id)

    response = client.get(
        f"/entities/{setup['entity_id']}/reports/kdv-input/export",
        params={"from": "2026-05-01", "to": "2026-05-31"},
    )
    _assert_xlsx_export(response, header_text="Rate (%)")


def test_kdv_input_export_invalid_date_range(
    client: TestClient, kdv_export_setup
) -> None:
    response = client.get(
        f"/entities/{kdv_export_setup['entity_id']}/reports/kdv-input/export",
        params={"from": "2026-05-31", "to": "2026-05-01"},
    )
    assert response.status_code == 422


def test_period_comparison_export(
    db_session, client: TestClient, export_setup
) -> None:
    setup = export_setup
    post_cash_movement(
        db_session,
        setup["entity_id"],
        money_account_id=setup["drawer"].id,
        movement_date=date(2026, 1, 10),
        direction=CashMovementDirection.IN,
        amount_kurus=100_000,
        offset_account_id=setup["accounts"][SALES_REVENUE_CODE],
        description="Cash sales",
        actor_id=ACTOR_ID,
    )
    pos_posting.post_card_sales_batch(
        db_session,
        setup["entity_id"],
        sales_date=date(2026, 1, 12),
        gross_amount_kurus=50_000,
        description="Card sales",
        actor_id=ACTOR_ID,
    )

    response = client.get(
        f"/entities/{setup['entity_id']}/reports/period-comparison/export",
        params={"from": "2026-01-01", "to": "2026-01-31"},
    )
    _assert_xlsx_export(response, header_text="Metric")


def test_period_comparison_export_invalid_date_range(
    client: TestClient, export_setup
) -> None:
    response = client.get(
        f"/entities/{export_setup['entity_id']}/reports/period-comparison/export",
        params={"from": "2026-02-01", "to": "2026-01-01"},
    )
    assert response.status_code == 422


def test_exported_amounts_are_lira(
    db_session, client: TestClient, export_setup
) -> None:
    """Exports used to carry raw kuruş, so 1.000 ₺ of rent read as 100000 and
    no figure could be checked against a statement without dividing by hand
    (2026-07-29)."""
    setup = export_setup
    _post_rent_expense(
        db_session, setup, amount_kurus=100_000, expense_date=date(2026, 1, 16)
    )

    response = client.get(
        f"/entities/{setup['entity_id']}/reports/profit-and-loss/export",
        params={"from": "2026-01-01", "to": "2026-01-31"},
    )
    ws = _load_sheet(response.content)
    numbers = [
        ws.cell(row=r, column=c)
        for r in range(1, ws.max_row + 1)
        for c in range(1, ws.max_column + 1)
        if isinstance(ws.cell(row=r, column=c).value, (int, float))
    ]
    values = [cell.value for cell in numbers]
    assert 1000.0 in values, "1.000,00 ₺ of rent should read as 1000.0"
    assert 100_000 not in values

    # A number, not a formatted string — otherwise every SUM() in the file breaks.
    # Money columns carry the accounting variant (negatives red in parentheses);
    # both are numeric formats, neither turns the value into text.
    assert all(
        cell.number_format in ("#,##0.00", "#,##0.00;[Red](#,##0.00)")
        for cell in numbers
    )


def test_money_columns_are_marked_as_lira(client: TestClient, export_setup) -> None:
    """Without the ₺ in the heading, a reader can't tell which unit they're
    looking at — the exact ambiguity that hid the kuruş bug."""
    response = client.get(
        f"/entities/{export_setup['entity_id']}/reports/profit-and-loss/export",
        params={"from": "2026-01-01", "to": "2026-01-31"},
    )
    ws = _load_sheet(response.content)
    text = " ".join(
        str(ws.cell(row=r, column=c).value)
        for r in range(1, ws.max_row + 1)
        for c in range(1, ws.max_column + 1)
    )
    assert "₺" in text


def test_every_exporter_still_imports() -> None:
    """These four modules share the workbook helpers, and renaming one helper
    broke three of them at import time without any test noticing — the export
    endpoints would have 500'd on first use (2026-07-29)."""
    from app.features.delivery import excel_export as delivery_export
    from app.features.payables import activity_excel
    from app.features.pos import excel_export as pos_export
    from app.features.reports import month_pack

    for module in (delivery_export, activity_excel, pos_export, month_pack):
        assert module is not None


def test_filename_slug_handles_turkish_names() -> None:
    """Restaurant names are Turkish; filenames have to be ascii.

    "İ".lower() is "i" followed by a combining dot in Python, so relying on
    lower() alone would put a stray mark in the filename. Each Turkish letter
    is mapped explicitly instead.
    """
    from app.features.reports.excel_export import filename_slug

    assert filename_slug("India Gate") == "india-gate"
    assert filename_slug("İstanbul Şubesi") == "istanbul-subesi"
    assert filename_slug("Çiğköfte & Co") == "cigkofte-co"
    assert filename_slug("Ali  Veli") == "ali-veli"
    # Nothing usable — callers leave the segment out rather than emit a leading hyphen.
    assert filename_slug("＊＊＊") == ""


def test_filename_slug_truncates_long_names_at_a_word_boundary() -> None:
    """Turkish trade names run long and must not be chopped mid-word.

    "MEHMET ÖZKAN GIDA SANAYİ VE TİCARET LİMİTED ŞİRKETİ" slugs to 51
    characters. A hard slice produced "...ticaret-limi" and a 82-character
    filename; cutting at the last hyphen gives what a person would have
    shortened it to.
    """
    from app.features.reports.excel_export import filename_slug

    assert (
        filename_slug("MEHMET ÖZKAN GIDA SANAYİ VE TİCARET LİMİTED ŞİRKETİ")
        == "mehmet-ozkan-gida"
    )
    # A dangling conjunction is dropped — "zaina-turizm-ve" reads as damaged.
    assert (
        filename_slug("ZAİNA TURİZM VE ORGANİZASYON HİZMETLERİ LTD ŞTİ")
        == "zaina-turizm"
    )
    # A real three-letter word at the end is kept.
    assert (
        filename_slug("ANADOLU ET VE SÜT ÜRÜNLERİ PAZARLAMA A.Ş.")
        == "anadolu-et-ve-sut"
    )
    # Short enough to survive whole.
    assert filename_slug("Metro Gastro") == "metro-gastro"
    # No word boundary to cut at — a hard slice is the only option.
    assert filename_slug("Supercalifragilisticexpialidocious") == (
        "supercalifragilisticexpi"
    )
    # Never ends on a hyphen, whatever the input.
    for name in ("A very long supplier name that keeps going and going", "x-" * 40):
        assert not filename_slug(name).endswith("-")


def test_export_filename_names_the_restaurant() -> None:
    """Two restaurants must not produce the same download.

    Without the entity segment every balance sheet arrives as
    balance-sheet-<date>, and India Gate's collides with Spice
    Corner's in the Downloads folder with nothing to tell them apart.
    """
    from datetime import date as date_cls

    from app.features.reports.excel_export import export_filename

    as_of = date_cls(2026, 6, 30)
    assert (
        export_filename("balance-sheet", entity_name="India Gate", as_of=as_of)
        == "india-gate-balance-sheet-2026-06-30.xlsx"
    )
    assert (
        export_filename("balance-sheet", entity_name="Spice Corner", as_of=as_of)
        != export_filename("balance-sheet", entity_name="India Gate", as_of=as_of)
    )
    # Dates stay ISO so filenames sort chronologically.
    assert "2026-06-30" in export_filename(
        "balance-sheet", entity_name="India Gate", as_of=as_of
    )
    # No entity given: no leading hyphen.
    assert (
        export_filename("balance-sheet", as_of=as_of)
        == "balance-sheet-2026-06-30.xlsx"
    )


def test_a_whole_month_is_written_as_one_segment() -> None:
    """2026-06 says what 2026-06-01-2026-06-30 says, in fourteen fewer
    characters, and most exports are whole months. A partial range has no
    shorter form, so it keeps both dates."""
    from datetime import date as date_cls

    from app.features.reports.excel_export import export_filename, period_segment

    assert period_segment(date_cls(2026, 6, 1), date_cls(2026, 6, 30)) == "2026-06"
    # Leap-aware: February ends on the 28th here, and that is still whole.
    assert period_segment(date_cls(2026, 2, 1), date_cls(2026, 2, 28)) == "2026-02"
    assert period_segment(date_cls(2024, 2, 1), date_cls(2024, 2, 29)) == "2024-02"
    # Not a whole month — both dates survive.
    assert (
        period_segment(date_cls(2026, 6, 10), date_cls(2026, 6, 20))
        == "2026-06-10-2026-06-20"
    )
    assert (
        period_segment(date_cls(2026, 2, 1), date_cls(2026, 2, 27))
        == "2026-02-01-2026-02-27"
    )

    assert (
        export_filename(
            "profit-and-loss",
            entity_name="India Gate",
            from_date=date_cls(2026, 6, 1),
            to_date=date_cls(2026, 6, 30),
        )
        == "india-gate-profit-and-loss-2026-06.xlsx"
    )
