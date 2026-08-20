"""Standalone cash-book / hand-recorded expenses / GL Excel downloads (S9)."""

from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from app.core.auth.types import EntityRole
from app.core.chart_of_accounts.seed import seed_default_chart
from app.core.chart_of_accounts.models import Account
from app.core.dates import format_period
from app.core.excel.labels import format_journal_source
from app.core.excel.workbook import money_header
from app.core.ledger.balances import balance_as_of_kurus
from app.core.ledger.models import JournalEntrySource, JournalEntryStatus
from app.core.period_locks.models import PeriodLockKind
from app.core.period_locks.service import close_period
from app.db.session import entity_context
from app.features.auth import service as auth_service
from app.features.auth.schema import MembershipCreate, UserCreate
from app.features.banking import service as banking_service
from app.features.banking.models import MoneyAccountKind
from app.features.banking.schema import MoneyAccountCreate
from app.features.expenses import service as expenses_service
from app.features.expenses.schema import ExpenseCreate
from app.features.reports import cash_book as cash_book_report
from app.features.reports import excel_export
from app.features.reports.cash_bank_book_export import cash_bank_book_filename
from tests.delivery_helpers import ACTOR_ID
from tests.test_financial_statements import (
    PERIOD_END,
    PERIOD_START,
    _post_period_sales,
    _post_rent_expense,
)

XLSX = excel_export.XLSX_CONTENT_TYPE


@pytest.fixture
def books_setup(db_session, restaurant_a):
    seed_default_chart(db_session, restaurant_a.id)
    owner = auth_service.create_user(
        db_session,
        UserCreate(email="books-export-owner@example.com", display_name="Owner"),
    )
    auth_service.add_entity_member(
        db_session,
        restaurant_a.id,
        MembershipCreate(user_id=owner.id, role=EntityRole.OWNER),
    )
    drawer = banking_service.create_money_account(
        db_session,
        restaurant_a.id,
        MoneyAccountCreate(account_kind=MoneyAccountKind.CASH, name="Main Drawer"),
    )
    bank = banking_service.create_money_account(
        db_session,
        restaurant_a.id,
        MoneyAccountCreate(
            account_kind=MoneyAccountKind.BANK, name="Garanti", bank_name="Garanti"
        ),
    )
    with entity_context(db_session, restaurant_a.id):
        accounts = {a.code: a.id for a in db_session.scalars(select(Account))}
    return {
        "entity_id": restaurant_a.id,
        "owner_id": owner.id,
        "drawer": drawer,
        "bank": bank,
        "accounts": accounts,
    }


def test_cash_bank_book_export_sheets_and_closing(
    db_session, client: TestClient, books_setup
) -> None:
    setup = books_setup
    _post_period_sales(db_session, setup)
    _post_rent_expense(
        db_session, setup, amount_kurus=20_000, expense_date=date(2026, 1, 16)
    )

    response = client.get(
        f"/entities/{setup['entity_id']}/reports/cash-book/export",
        params={"from": "2026-01-01", "to": "2026-01-31"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == XLSX
    disposition = response.headers.get("content-disposition", "")
    assert "mizan-cash-bank-book-2026-01-" in disposition
    assert disposition.endswith('.xlsx"') or ".xlsx" in disposition

    wb = load_workbook(BytesIO(response.content))
    names = wb.sheetnames
    assert any(n.startswith("Cash — ") for n in names)
    assert any(n.startswith("Bank — ") for n in names)

    # Amounts are lira floats, not raw kuruş.
    cash_ws = next(wb[n] for n in names if n.startswith("Cash — "))
    found_lira = False
    for row in cash_ws.iter_rows(min_row=1, max_row=cash_ws.max_row, max_col=6):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value == 1000.0:
                found_lira = True
            assert cell.value != 100_000
    assert found_lira

    book = cash_book_report.get_cash_book(
        db_session,
        setup["entity_id"],
        setup["drawer"].id,
        PERIOD_START,
        PERIOD_END,
    )
    # Closing cell is row 4 col 2 in the month-pack account-book layout.
    closing_cell = cash_ws.cell(row=4, column=2).value
    assert closing_cell == pytest.approx(book.closing_kurus / 100, abs=0.01)

    with entity_context(db_session, setup["entity_id"]):
        gl = db_session.get(Account, setup["drawer"].gl_account_id)
        gl_closing = balance_as_of_kurus(db_session, gl, PERIOD_END)
    assert book.closing_kurus == gl_closing


def test_cash_bank_book_quiet_account_still_gets_a_sheet(
    db_session, client: TestClient, books_setup
) -> None:
    """Account with no movements in range: sheet still present, opening = closing."""
    setup = books_setup
    # Activity only on the cash drawer — bank stays quiet for January.
    _post_period_sales(db_session, setup)

    response = client.get(
        f"/entities/{setup['entity_id']}/reports/cash-book/export",
        params={"from": "2026-01-01", "to": "2026-01-31"},
    )
    assert response.status_code == 200
    wb = load_workbook(BytesIO(response.content))
    bank_ws = next(
        wb[n] for n in wb.sheetnames if n.startswith("Bank — ")
    )

    book = cash_book_report.get_cash_book(
        db_session,
        setup["entity_id"],
        setup["bank"].id,
        PERIOD_START,
        PERIOD_END,
    )
    assert book.rows == []
    opening = bank_ws.cell(row=3, column=2).value
    closing = bank_ws.cell(row=4, column=2).value
    assert opening == pytest.approx(book.opening_kurus / 100, abs=0.01)
    assert closing == pytest.approx(book.closing_kurus / 100, abs=0.01)
    assert opening == closing
    # No movement dates under the header row.
    assert bank_ws.cell(row=7, column=1).value is None


def test_cash_bank_book_opening_and_running_balance(
    db_session, client: TestClient, books_setup
) -> None:
    """Opening is day-before-From; each movement row carries a running Balance."""
    setup = books_setup
    _post_period_sales(db_session, setup)

    response = client.get(
        f"/entities/{setup['entity_id']}/reports/cash-book/export",
        params={"from": "2026-01-01", "to": "2026-01-31"},
    )
    assert response.status_code == 200
    wb = load_workbook(BytesIO(response.content))
    cash_ws = next(wb[n] for n in wb.sheetnames if n.startswith("Cash — "))

    assert cash_ws.cell(row=3, column=1).value == "Opening"
    headers = [cash_ws.cell(row=6, column=c).value for c in range(1, 7)]
    assert headers == [
        "Date",
        "Description",
        "Recorded as",
        money_header("In"),
        money_header("Out"),
        money_header("Balance"),
    ]

    book = cash_book_report.get_cash_book(
        db_session,
        setup["entity_id"],
        setup["drawer"].id,
        PERIOD_START,
        PERIOD_END,
    )
    with entity_context(db_session, setup["entity_id"]):
        gl = db_session.get(Account, setup["drawer"].gl_account_id)
        day_before = balance_as_of_kurus(
            db_session, gl, PERIOD_START - timedelta(days=1)
        )
    assert book.opening_kurus == day_before
    assert cash_ws.cell(row=3, column=2).value == pytest.approx(
        day_before / 100, abs=0.01
    )
    assert book.rows, "expected cash movements so running balances can be checked"
    for index, line in enumerate(book.rows):
        row = 7 + index
        assert cash_ws.cell(row=row, column=6).value == pytest.approx(
            line.balance_kurus / 100, abs=0.01
        )


def test_cash_bank_book_filename_suffix() -> None:
    assert cash_bank_book_filename(
        date(2026, 1, 1), date(2026, 1, 31), sealed=False
    ).endswith("-live.xlsx")
    assert cash_bank_book_filename(
        date(2026, 1, 1), date(2026, 1, 31), sealed=True
    ).endswith("-as-closed.xlsx")


def test_hand_recorded_expenses_export_total_matches_list(
    db_session, client: TestClient, books_setup
) -> None:
    setup = books_setup
    expenses_service.create_expense(
        db_session,
        setup["entity_id"],
        ExpenseCreate(
            expense_date=date(2026, 1, 10),
            amount_kurus=15_000,
            expense_account_id=setup["accounts"]["5200"],
            money_account_id=setup["drawer"].id,
            written_item_description="Tea towels",
            description="Tea towels",
            notes="kitchen",
            actor_id=ACTOR_ID,
        ),
    )
    expenses_service.create_expense(
        db_session,
        setup["entity_id"],
        ExpenseCreate(
            expense_date=date(2026, 1, 12),
            amount_kurus=25_000,
            expense_account_id=setup["accounts"]["5200"],
            money_account_id=setup["drawer"].id,
            written_item_description="Soap",
            description="Soap",
            actor_id=ACTOR_ID,
        ),
    )

    list_res = client.get(
        f"/entities/{setup['entity_id']}/expenses",
        params={"from": "2026-01-01", "to": "2026-01-31", "limit": 200},
    )
    assert list_res.status_code == 200
    page_total = list_res.json()["total_amount_kurus"]
    assert page_total == 40_000

    response = client.get(
        f"/entities/{setup['entity_id']}/expenses/export",
        params={"from": "2026-01-01", "to": "2026-01-31"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == XLSX
    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    assert ws["A1"].value == "Hand-recorded expenses"
    headers = [ws.cell(row=4, column=c).value for c in range(1, 7)]
    assert headers == [
        "Date",
        "Item",
        "Account",
        "Paid from",
        "Note",
        money_header("Amount"),
    ]

    # Footer total (last money cell in column 6).
    total_cell = None
    for r in range(ws.max_row, 0, -1):
        if ws.cell(row=r, column=5).value == "Total":
            total_cell = ws.cell(row=r, column=6).value
            break
    assert total_cell == pytest.approx(page_total / 100, abs=0.01)


def test_general_ledger_export_sheets_and_tie(
    db_session, client: TestClient, books_setup
) -> None:
    setup = books_setup
    _post_period_sales(db_session, setup)
    _post_rent_expense(
        db_session, setup, amount_kurus=20_000, expense_date=date(2026, 1, 16)
    )

    list_res = client.get(
        f"/entities/{setup['entity_id']}/ledger/entries",
        params={
            "from": "2026-01-01",
            "to": "2026-01-31",
            "effective_only": "true",
            "limit": 200,
        },
    )
    assert list_res.status_code == 200
    entries = list_res.json()["items"]
    expected_lines = sum(len(e["lines"]) for e in entries)

    response = client.get(
        f"/entities/{setup['entity_id']}/ledger/entries/export",
        params={
            "from": "2026-01-01",
            "to": "2026-01-31",
            "effective_only": "true",
        },
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == XLSX
    disposition = response.headers.get("content-disposition", "")
    assert "-live.xlsx" in disposition or "-as-closed.xlsx" in disposition

    wb = load_workbook(BytesIO(response.content))
    assert wb.sheetnames == ["By account", "All entries"]

    by_acct = wb["By account"]
    # Data starts at row 5; each row must tie opening + signed period = closing.
    for r in range(5, by_acct.max_row + 1):
        code = by_acct.cell(row=r, column=1).value
        if not code:
            continue
        opening = by_acct.cell(row=r, column=3).value or 0
        debits = by_acct.cell(row=r, column=4).value or 0
        credits = by_acct.cell(row=r, column=5).value or 0
        closing = by_acct.cell(row=r, column=6).value or 0
        # Compare in lira space with debit-normal identity for cash/expense
        # and credit-normal for revenue — recompute via GL for the code.
        with entity_context(db_session, setup["entity_id"]):
            account = db_session.scalar(
                select(Account).where(Account.code == code)
            )
            assert account is not None
            from app.core.chart_of_accounts.types import AccountNormalBalance

            if account.normal_balance == AccountNormalBalance.DEBIT:
                assert closing == pytest.approx(opening + debits - credits, abs=0.01)
            else:
                assert closing == pytest.approx(opening + credits - debits, abs=0.01)

    detail = wb["All entries"]
    detail_headers = [detail.cell(row=4, column=c).value for c in range(1, 7)]
    assert detail_headers == [
        "Date",
        "Description",
        "Source",
        "Account",
        money_header("Debit"),
        money_header("Credit"),
    ]
    data_lines = 0
    for r in range(5, detail.max_row + 1):
        if detail.cell(row=r, column=1).value:
            data_lines += 1
    assert data_lines == expected_lines


def test_general_ledger_export_stamps_range_and_filters(
    db_session, client: TestClient, books_setup
) -> None:
    """Header carries the applied from/to and any active source/status filter."""
    setup = books_setup
    _post_period_sales(db_session, setup)

    response = client.get(
        f"/entities/{setup['entity_id']}/ledger/entries/export",
        params={
            "from": "2026-01-01",
            "to": "2026-01-31",
            "source": JournalEntrySource.CASH_MOVEMENT.value,
            "status": JournalEntryStatus.POSTED.value,
            "effective_only": "false",
        },
    )
    assert response.status_code == 200
    wb = load_workbook(BytesIO(response.content))
    period = format_period(PERIOD_START, PERIOD_END)
    source_label = format_journal_source(JournalEntrySource.CASH_MOVEMENT.value)
    for sheet_name in ("By account", "All entries"):
        ws = wb[sheet_name]
        subtitle_period = str(ws.cell(row=2, column=1).value)
        subtitle_filters = str(ws.cell(row=3, column=1).value)
        assert period in subtitle_period
        assert f"Source: {source_label}" in subtitle_filters
        assert "Status: posted" in subtitle_filters


def test_closed_month_exports_as_closed_filename(
    db_session, client: TestClient, books_setup
) -> None:
    """A sealed month stamps -as-closed on cash-book and GL downloads."""
    setup = books_setup
    _post_period_sales(db_session, setup)
    close_period(
        db_session,
        setup["entity_id"],
        lock_kind=PeriodLockKind.MONTH,
        anchor_date=PERIOD_END,
        actor_id=setup["owner_id"],
    )

    cash = client.get(
        f"/entities/{setup['entity_id']}/reports/cash-book/export",
        params={"from": "2026-01-01", "to": "2026-01-31"},
    )
    assert cash.status_code == 200
    assert "-as-closed.xlsx" in cash.headers.get("content-disposition", "")

    gl = client.get(
        f"/entities/{setup['entity_id']}/ledger/entries/export",
        params={"from": "2026-01-01", "to": "2026-01-31", "effective_only": "true"},
    )
    assert gl.status_code == 200
    assert "-as-closed.xlsx" in gl.headers.get("content-disposition", "")
