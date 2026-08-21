"""Partner ledger Excel/PDF download endpoints."""

from __future__ import annotations

import uuid
from datetime import date, datetime
from decimal import Decimal

from fastapi.testclient import TestClient

from app.core.chart_of_accounts.seed import seed_default_chart
from app.core.partners import profit_allocation as pa
from app.db.session import entity_context
from app.features.partners.models import Partner


ACTOR_ID = uuid.UUID("00000000-0000-4000-8000-000000000001")


def _partner_setup(db_session, restaurant_a):
    seed_default_chart(db_session, restaurant_a.id)
    with entity_context(db_session, restaurant_a.id):
        partners = []
        for name, pct in [("Ali", "60"), ("Burak", "40")]:
            p = Partner(name=name, ownership_share_pct=Decimal(pct))
            db_session.add(p)
            partners.append(p)
        db_session.commit()
        for p in partners:
            db_session.refresh(p)
    return restaurant_a.id, partners[0]


def test_partner_ledger_export_xlsx_and_pdf(
    db_session, restaurant_a, client: TestClient
) -> None:
    entity_id, partner = _partner_setup(db_session, restaurant_a)
    partner_id = partner.id
    pa.post_profit_allocation(
        db_session,
        entity_id,
        allocation_date=date(2026, 6, 30),
        profit_kurus=100_000,
        description="Export sample",
        actor_id=ACTOR_ID,
        net_against_drawings=False,
        netting_as_of=date(2026, 6, 30),
    )

    xlsx = client.get(
        f"/entities/{entity_id}/partners/{partner_id}/ledger/export"
    )
    assert xlsx.status_code == 200, xlsx.text
    assert "spreadsheetml" in xlsx.headers["content-type"]
    assert xlsx.content[:2] == b"PK"

    pdf = client.get(
        f"/entities/{entity_id}/partners/{partner_id}/ledger/export/pdf"
    )
    assert pdf.status_code == 200, pdf.text
    assert pdf.headers["content-type"].startswith("application/pdf")
    assert pdf.content[:4] == b"%PDF"


def _movement_rows(ws):
    """Rows below the header, found by the header rather than by counting.

    This read from a hard-coded `min_row=12`, which was the header's position
    when the summary carried seven figures. Trimming it to three moved every
    row up four and the test read past the end of the sheet — a failure about
    the summary's height, reported as "no movements". The sheet says where its
    own table starts.
    """
    header = next(
        row[0].row for row in ws.iter_rows(min_col=1, max_col=1) if row[0].value == "Date"
    )
    return list(ws.iter_rows(min_row=header + 1))


def test_partner_ledger_download_shows_the_ledger_as_it_now_stands(
    db_session, restaurant_a, client: TestClient
) -> None:
    """A corrected allocation downloads as one row, not three.

    Correcting voids the original and reposts, so `get_partner_ledger` holds
    three rows: the superseded original, the `Void: …` reversal, and the
    replacement. The screen hides the first two behind a history toggle. A
    download has no toggle, and the PDF has no status column either — so an
    unfiltered export reads as three real movements, two of which never
    happened.
    """
    import io

    from openpyxl import load_workbook

    entity_id, partner = _partner_setup(db_session, restaurant_a)
    partner_id = partner.id
    posted = pa.post_profit_allocation(
        db_session,
        entity_id,
        allocation_date=date(2026, 6, 30),
        profit_kurus=100_000,
        description="First figure",
        actor_id=ACTOR_ID,
        net_against_drawings=False,
        netting_as_of=date(2026, 6, 30),
    )
    pa.correct_profit_allocation(
        db_session,
        entity_id,
        posted.journal_entry.id,
        allocation_date=date(2026, 6, 30),
        profit_kurus=250_000,
        description="Corrected figure",
        actor_id=ACTOR_ID,
        net_against_drawings=False,
        netting_as_of=date(2026, 6, 30),
    )

    resp = client.get(f"/entities/{entity_id}/partners/{partner_id}/ledger/export")
    assert resp.status_code == 200, resp.text
    ws = load_workbook(io.BytesIO(resp.content)).active
    descriptions = [
        str(row[2].value) for row in _movement_rows(ws) if row[2].value is not None
    ]

    assert descriptions, "expected at least the surviving allocation row"
    assert not any(d.startswith("Void:") for d in descriptions), descriptions
    assert not any("First figure" in d for d in descriptions), descriptions
    assert any("Corrected figure" in d for d in descriptions), descriptions

    # Dates are real date cells, not text. Written as a string, Excel sorts
    # them alphabetically and cannot filter by month.
    dates = [row[0].value for row in _movement_rows(ws) if row[0].value]
    assert dates, "expected a date on the surviving row"
    assert all(isinstance(value, (date, datetime)) for value in dates), dates


def _pdf_text(data: bytes) -> str:
    import io

    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    return "".join(page.extract_text() or "" for page in reader.pages)


def test_partner_ledger_pdf_uses_the_shared_report_furniture(
    db_session, restaurant_a, client: TestClient
) -> None:
    """Masthead, KPI strip and footer — not a bare table.

    This export used to hand-roll its own title and a full-grid table, so it
    came out looking like a printout rather than a statement next to the P&L
    and balance sheet. It now builds from header_elements / summary_band /
    _table_style like they do.
    """
    entity_id, partner = _partner_setup(db_session, restaurant_a)
    # Read off the instance before anything commits. Posting the allocation
    # expires it, and touching an attribute afterwards makes SQLAlchemy refresh
    # a row this session can no longer see — ObjectDeletedError, not a real
    # deletion. The other tests in this file capture the id the same way.
    partner_id = partner.id
    partner_name = partner.name
    pa.post_profit_allocation(
        db_session,
        entity_id,
        allocation_date=date(2026, 6, 30),
        profit_kurus=100_000,
        description="Export sample",
        actor_id=ACTOR_ID,
        net_against_drawings=False,
        netting_as_of=date(2026, 6, 30),
    )

    resp = client.get(
        f"/entities/{entity_id}/partners/{partner_id}/ledger/export/pdf"
    )
    assert resp.status_code == 200, resp.text
    # Whitespace-normalised: pypdf's line breaks depend on the layout, and
    # this test is about what the page says, not where it wraps.
    text = " ".join(_pdf_text(resp.content).split())

    # Masthead: titled, attributed, dated.
    assert "Partner ledger" in text
    assert partner_name in text
    assert "As at" in text
    # KPI strip, in the partner page's order. Three lines now, not seven: the
    # statement prints a figure only when the partner has one. "Unpaid profit"
    # and "settled from drawings" went with the summary cards — the balance
    # already nets the profit, and the allocation rows show the split.
    for label in ("NET BALANCE", "CAPITAL IN BUSINESS", "PROFIT ALLOCATED"):
        assert label in text, f"missing KPI {label!r}"
    for gone in ("UNPAID PROFIT", "SETTLED FROM DRAWINGS"):
        assert gone not in text, f"{gone!r} is still on the statement"
    # Footer stamp that every other report PDF carries.
    assert "Mizan" in text
    # Dates read the way the app shows them everywhere else, not ISO.
    assert "30.06.2026" in text
    assert "2026-06-30" not in text


def test_partner_ledger_pdf_wraps_long_descriptions(
    db_session, restaurant_a, client: TestClient
) -> None:
    """A bank reference has to wrap inside its column, not run over the money.

    Descriptions carry whole payment references — IBANs, SGK numbers,
    counterparty names. A plain string in a reportlab table cell does not
    wrap; it overflows across the Amount and Running columns, which is what
    this export used to do. The old guard against that was truncating to 80
    characters, which mangled the reference without stopping the overflow.
    """
    entity_id, partner = _partner_setup(db_session, restaurant_a)
    partner_id = partner.id
    reference = (
        "SYED FAIZAN ALI BUKHARI*TR470006400000175030614324*MASALY*"
        "H2606620775209 Sicil: 25610010110437550500152000 Borc Kodu: 04101"
    )
    pa.post_profit_allocation(
        db_session,
        entity_id,
        allocation_date=date(2026, 6, 30),
        profit_kurus=100_000,
        description=reference,
        actor_id=ACTOR_ID,
        net_against_drawings=False,
        netting_as_of=date(2026, 6, 30),
    )

    resp = client.get(
        f"/entities/{entity_id}/partners/{partner_id}/ledger/export/pdf"
    )
    assert resp.status_code == 200, resp.text

    # Whitespace stripped entirely, not just normalised. A wrapped line is the
    # point of this test, and pypdf reports the break as an extra space
    # mid-reference ("…*MASALY *H2606620775209"), so any comparison that keeps
    # spaces fails on precisely the behaviour being asserted.
    squashed = "".join(_pdf_text(resp.content).split())

    # Present in full — wrapping keeps it, truncation would have cut it at 80.
    assert len(reference) > 80, "fixture must exceed the old truncation point"
    assert "".join(reference.split()) in squashed, (
        "description was truncated or dropped"
    )


def test_partner_ledger_export_missing_partner_404(
    db_session, restaurant_a, client: TestClient
) -> None:
    entity_id = restaurant_a.id
    seed_default_chart(db_session, entity_id)
    missing = uuid.uuid4()
    resp = client.get(
        f"/entities/{entity_id}/partners/{missing}/ledger/export"
    )
    assert resp.status_code == 404


class TestTheSummaryPrintsOnlyWhatIsTrue:
    """Seven lines, four of them zero for most partners.

    "The stickers are explaining too much" was said of the page; the statement
    had the same fault. A document that prints every term it knows makes the
    reader find the two figures that moved.
    """

    def _summary(self, **figures):
        from app.features.partners import ledger_export
        from app.features.partners.schema import PartnerLedgerRead

        # `entries` has no default; the summary never reads it.
        return ledger_export._summary(
            PartnerLedgerRead(partner_id=uuid.uuid4(), entries=[], **figures)
        )

    def test_the_balance_prints_even_at_zero(self) -> None:
        # Zero is an answer. "Settled" is worth saying out loud.
        labels = [label for label, _ in self._summary(balance_kurus=0)]
        assert labels == ["Net balance"]

    def test_a_figure_the_partner_does_not_have_is_left_out(self) -> None:
        labels = [
            label
            for label, _ in self._summary(
                balance_kurus=0,
                current_account_kurus=-1_203_609,
                capital_balance_kurus=53_246_391,
                profit_allocated_kurus=17_500_000,
            )
        ]
        assert labels == ["Net balance", "Capital in business", "Profit allocated"]

    def test_a_loan_or_fronted_expense_earns_its_line(self) -> None:
        # Guard the guard: if the conditions were inverted, the test above
        # would still pass while these two never appeared at all.
        labels = [
            label
            for label, _ in self._summary(
                balance_kurus=45_000,
                loan_balance_kurus=900_000,
            )
        ]
        assert "Fronted expenses" in labels
        assert "Partner loan" in labels

    def test_the_balance_is_the_netted_one(self) -> None:
        # Not `net_balance_kurus`, which excludes profit already credited and
        # exists to decide settlement rather than to be read.
        summary = dict(
            self._summary(balance_kurus=0, current_account_kurus=-1_203_609)
        )
        assert summary["Net balance"] == -1_203_609


_ALIGN_TOL_PT = 2.0


def _header_word(words, page_index: int, label: str):
    matches = [
        w for pi, w in words if pi == page_index and w[4] == label
    ]
    assert matches, f"{label!r} header missing on page {page_index + 1}"
    return min(matches, key=lambda w: w[1])


def _money_tail_right(words, page_index: int, *, below_y: float, above_y: float, col_left: float, col_right: float) -> float:
    """Right edge of the trailing ₺ (or last money fragment) in a column band."""
    tails = [
        w
        for pi, w in words
        if pi == page_index
        and below_y < w[1] < above_y
        and col_left <= (w[0] + w[2]) / 2 <= col_right
        and (w[4] == "₺" or "," in w[4])
    ]
    assert tails, "expected money in column band"
    return max(w[2] for w in tails)


def test_subledger_pdf_money_columns_share_one_geometry() -> None:
    """Amount / Running header right edges match value right edges — every page.

    Owner report (India Gate / Canan Takan, 2026-08-21): headers sat left in
    the money columns while figures sat right. Geometry is defined once and
    used for the header, body, and repeated page-2 header; long wrapping
    descriptions must not shove the money columns sideways.
    """
    from app.features.reports.subledger_export import (
        SubledgerExport,
        SubledgerRow,
        build_subledger_pdf,
    )

    long_desc = (
        "SYED FAIZAN ALI BUKHARI*TR470006400000175030614324*MASALY*"
        "H2606620775209 Sicil: 25610010110437550500152000 Borc Kodu: 04101"
    )
    rows = [
        SubledgerRow(
            movement_date=date(2026, 6, 1),
            movement="Profit allocation",
            description="Short",
            amount_minor=100_000,
            running_minor=100_000,
        ),
        SubledgerRow(
            movement_date=date(2026, 6, 2),
            movement="Drawing",
            description=long_desc,
            amount_minor=-50_000,
            running_minor=50_000,
        ),
    ]
    for i in range(40):
        rows.append(
            SubledgerRow(
                movement_date=date(2026, 6, 3),
                movement="Profit allocation",
                description=f"pad-{i}",
                amount_minor=10_000,
                running_minor=60_000 + i * 10_000,
            )
        )

    data = build_subledger_pdf(
        SubledgerExport(
            entity_name="India Gate",
            subject_name="Canan Takan",
            ledger_label="Partner ledger",
            sheet_name="Partner",
            summary=[
                ("Net balance", 1_234_567),
                ("Capital in business", 5_000_000),
                ("Profit allocated", 2_000_000),
                ("Fronted expenses", 900_000),
            ],
            rows=rows,
        )
    )

    import fitz

    with fitz.open(stream=data, filetype="pdf") as doc:
        assert doc.page_count >= 2, "fixture must span a repeated header page"
        words = [
            (page_index, word)
            for page_index, page in enumerate(doc)
            for word in page.get_text("words")
        ]

    # Page-1 and page-2 repeated headers share the same right edges.
    for label in ("Amount", "Running"):
        p1 = _header_word(words, 0, label)
        p2 = _header_word(words, 1, label)
        assert abs(p1[2] - p2[2]) <= _ALIGN_TOL_PT, (
            f"{label} header right page1={p1[2]:.2f} page2={p2[2]:.2f}"
        )

    amount_hdr = _header_word(words, 0, "Amount")
    running_hdr = _header_word(words, 0, "Running")
    # Column bands: midpoints fall inside the money columns.
    amount_band = (amount_hdr[0] - 40, (amount_hdr[2] + running_hdr[0]) / 2)
    running_band = ((amount_hdr[2] + running_hdr[0]) / 2, running_hdr[2] + 40)

    # Short-description row (page 1).
    short = next(w for pi, w in words if pi == 0 and w[4] == "Short")
    short_amt_right = _money_tail_right(
        words,
        0,
        below_y=short[1] - 2,
        above_y=short[3] + 2,
        col_left=amount_band[0],
        col_right=amount_band[1],
    )
    short_run_right = _money_tail_right(
        words,
        0,
        below_y=short[1] - 2,
        above_y=short[3] + 2,
        col_left=running_band[0],
        col_right=running_band[1],
    )
    assert abs(short_amt_right - amount_hdr[2]) <= _ALIGN_TOL_PT, (
        f"Amount header {amount_hdr[2]:.2f} vs value {short_amt_right:.2f}"
    )
    assert abs(short_run_right - running_hdr[2]) <= _ALIGN_TOL_PT, (
        f"Running header {running_hdr[2]:.2f} vs value {short_run_right:.2f}"
    )

    # Long wrapping description — money columns keep the same right edge.
    long_frag = next(
        w for pi, w in words if pi == 0 and w[4].startswith("SYED")
    )
    long_words_y0 = long_frag[1]
    long_words_y1 = max(
        w[3]
        for pi, w in words
        if pi == 0 and (w[4].startswith("SYED") or "TR4700" in w[4] or "Sicil" in w[4])
    )
    long_amt_right = _money_tail_right(
        words,
        0,
        below_y=long_words_y0 - 2,
        above_y=long_words_y1 + 4,
        col_left=amount_band[0],
        col_right=amount_band[1],
    )
    long_run_right = _money_tail_right(
        words,
        0,
        below_y=long_words_y0 - 2,
        above_y=long_words_y1 + 4,
        col_left=running_band[0],
        col_right=running_band[1],
    )
    assert abs(long_amt_right - short_amt_right) <= _ALIGN_TOL_PT
    assert abs(long_run_right - short_run_right) <= _ALIGN_TOL_PT

    # Four-box summary: each label shares its box's left edge with its value.
    net = next(w for pi, w in words if pi == 0 and w[4] == "NET")
    value = next(w for pi, w in words if pi == 0 and w[4] == "12.345,67")
    assert abs(net[0] - value[0]) <= _ALIGN_TOL_PT
