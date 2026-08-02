"""Shared openpyxl helpers for report export."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook as WorkbookType
from openpyxl.worksheet.worksheet import Worksheet

# Design-system palette (DESIGN_SYSTEM.md) — blue accent, calm greens/reds.
_BLUE = "2563EB"
_BLUE_DARK = "1D4ED8"
_BLUE_SOFT = "DBEAFE"
_GREEN = "16A34A"
_GREEN_SOFT = "DCFCE7"
_RED = "DC2626"
_RED_SOFT = "FEE2E2"
_SLATE = "334155"
_MUTED = "64748B"

_TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=_BLUE_DARK)
_SUBTITLE_FONT = Font(name="Calibri", size=10, color=_MUTED)
_META_LABEL_FONT = Font(name="Calibri", size=10, bold=True, color=_SLATE)
_HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor=_BLUE)
_SECTION_FILL = PatternFill("solid", fgColor=_BLUE_SOFT)
_SECTION_FONT = Font(name="Calibri", size=10, bold=True, color=_BLUE_DARK)
_STRIPE_FILL = PatternFill("solid", fgColor="F0F9FF")
_OPENING_FILL = PatternFill("solid", fgColor=_BLUE_SOFT)
_CLOSING_FILL = PatternFill("solid", fgColor=_GREEN_SOFT)
_TOTAL_FILL = PatternFill("solid", fgColor=_BLUE_SOFT)
_THIN = Side(style="thin", color="BFDBFE")
_HEADER_BORDER = Border(bottom=Side(style="medium", color=_BLUE_DARK))
_CELL_BORDER = Border(
    left=_THIN, right=_THIN, top=_THIN, bottom=_THIN
)
_BOLD = Font(name="Calibri", bold=True)
_DATE_FORMAT = "DD.MM.YYYY"

# Public tokens for report sheets that need opening / closing / signed tints.
OPENING_FILL = _OPENING_FILL
CLOSING_FILL = _CLOSING_FILL
TOTAL_FILL = _TOTAL_FILL
LOSS_FILL = PatternFill("solid", fgColor=_RED_SOFT)
BLUE_DARK = _BLUE_DARK
GREEN = _GREEN
RED = _RED
SUBTITLE_FONT = _SUBTITLE_FONT


def create_workbook(sheet_title: str = "Report") -> tuple[Workbook, Worksheet]:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = unique_sheet_title(wb, sheet_title, replace_active=True)
    return wb, ws


def unique_sheet_title(
    wb: WorkbookType, base: str, *, replace_active: bool = False
) -> str:
    """Excel caps titles at 31 characters and rejects duplicates — return a safe unique name."""
    cleaned = " ".join((base or "Sheet").split())
    if not cleaned:
        cleaned = "Sheet"
    # Excel forbids: : \\ / ? * [ ]
    for ch in (":", "\\", "/", "?", "*", "[", "]"):
        cleaned = cleaned.replace(ch, "-")
    stem = cleaned[:31]
    existing = {s.title for s in wb.worksheets}
    if replace_active and wb.active is not None:
        # Active sheet is being renamed; ignore its current title for uniqueness.
        existing.discard(wb.active.title)
    if stem not in existing:
        return stem
    # Truncate stem so " (2)" etc. still fits in 31.
    for n in range(2, 100):
        suffix = f" ({n})"
        candidate = f"{stem[: 31 - len(suffix)]}{suffix}"
        if candidate not in existing:
            return candidate
    return f"{stem[:28]}_{len(existing)}"


def add_sheet(wb: WorkbookType, title: str) -> Worksheet:
    """Append a sheet with a collision-safe title (max 31 chars)."""
    return wb.create_sheet(title=unique_sheet_title(wb, title))


#: Plain numeric format. Deliberately not a literal like "#.##0,00" — Excel
#: renders this per the *reader's* locale, so a Turkish machine shows
#: 1.234,50 and an English one 1,234.50, both from the same file.
MONEY_FORMAT = "#,##0.00"


def money_header(column_name: str = "Amount") -> str:
    return f"{column_name} (₺)"


def quantity_header(currency: str, column_name: str = "Amount held") -> str:
    """Native FX quantity column — currency in the header, not the cell format."""
    code = (currency or "").strip().upper() or "FX"
    return f"{column_name} ({code})"


def write_money(ws, row: int, col: int, minor: int | None) -> None:
    """Write a minor-unit amount as lira, still a number Excel can sum.

    Amounts are stored in kuruş so the ledger never touches a float, but a
    column of raw kuruş is unreadable and can't be checked against a statement
    — the owner would have to divide every figure by 100 by hand
    (2026-07-29). Written as a real number, not text, so filters and SUM()
    keep working.
    """
    if minor is None:
        return
    cell = ws.cell(row=row, column=col, value=minor / 100)
    cell.number_format = MONEY_FORMAT


def write_quantity(ws, row: int, col: int, minor: int | None) -> None:
    """A foreign-currency quantity held in minor units — €, $, £, not ₺.

    Arithmetically identical to write_money today, and kept separate anyway:
    these two columns sit side by side on the foreign currency sheet and mean
    different things, so a future change to one (a ₺ symbol in the format, say)
    must not silently follow into the other.
    """
    write_money(ws, row, col, minor)


def bold_row(ws: Worksheet, row: int, *, start_col: int = 1, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(
            name=cell.font.name or "Calibri",
            size=cell.font.size or 11,
            bold=True,
            color=cell.font.color,
        )


def tint_row(
    ws: Worksheet,
    row: int,
    *,
    end_col: int,
    fill: PatternFill,
    font_color: str | None = None,
    bold: bool = False,
) -> None:
    """Soft row highlight for opening / closing / total lines."""
    for col in range(1, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = Font(
            name="Calibri",
            size=cell.font.size or 11,
            bold=bold or bool(cell.font.bold),
            color=font_color or _SLATE,
        )


def style_signed_money(ws: Worksheet, row: int, col: int, minor: int | None) -> None:
    """Green for money in, red for money out — keeps the number Excel-summable."""
    if minor is None:
        return
    cell = ws.cell(row=row, column=col)
    if minor > 0:
        cell.font = Font(name="Calibri", color=_GREEN)
    elif minor < 0:
        cell.font = Font(name="Calibri", color=_RED)


def write_date(ws: Worksheet, row: int, col: int, value: date | datetime | None) -> None:
    if value is None:
        return
    if isinstance(value, datetime):
        value = value.date()
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = _DATE_FORMAT
    cell.alignment = Alignment(horizontal="center", vertical="center")


def write_meta_pair(ws: Worksheet, row: int, label: str, value: object) -> None:
    ws.cell(row=row, column=1, value=label).font = _META_LABEL_FONT
    value_cell = ws.cell(row=row, column=2, value=value)
    value_cell.font = _SUBTITLE_FONT


def write_section_header(
    ws: Worksheet, row: int, label: str, *, end_col: int = 2
) -> int:
    ws.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=end_col
    )
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = _SECTION_FONT
    cell.fill = _SECTION_FILL
    cell.border = Border(bottom=_THIN)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 18
    return row + 1


def set_column_widths(ws: Worksheet, widths: dict[int, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def style_money_columns(
    ws: Worksheet,
    *,
    money_cols: tuple[int, ...],
    first_row: int,
    last_row: int,
) -> None:
    for row in range(first_row, last_row + 1):
        for col in money_cols:
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            cell.alignment = Alignment(horizontal="right", vertical="center")


def write_sheet_title(
    ws: Worksheet,
    title: str,
    *,
    subtitles: list[str] | None = None,
    end_col: int = 2,
) -> int:
    """Title + optional subtitle lines. Returns the next free row."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24
    row = 2
    for line in subtitles or []:
        cell = ws.cell(row=row, column=1, value=line)
        cell.font = _SUBTITLE_FONT
        row += 1
    return row + 1  # blank spacer before content


def write_header_row(
    ws: Worksheet, row: int, headers: list[str], *, start_col: int = 1
) -> int:
    """Styled column headers. Returns the first data row."""
    for offset, header in enumerate(headers):
        col = start_col + offset
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _HEADER_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    return row + 1


def finish_data_table(
    ws: Worksheet,
    *,
    header_row: int,
    last_data_row: int,
    end_col: int,
    freeze_panes: str | None = None,
    autofilter: bool = True,
    money_cols: tuple[int, ...] | None = None,
    zebra: bool = True,
    column_widths: dict[int, float] | None = None,
) -> None:
    """Freeze header, optional AutoFilter, light borders, autosize."""
    if last_data_row >= header_row:
        for r in range(header_row + 1, last_data_row + 1):
            for c in range(1, end_col + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    continue
                left = cell.border.left if cell.border is not None else None
                if left is None or left.style is None:
                    cell.border = _CELL_BORDER
                if zebra and (r - header_row) % 2 == 0:
                    if cell.fill is None or cell.fill.fill_type is None:
                        cell.fill = _STRIPE_FILL
        if money_cols:
            style_money_columns(
                ws,
                money_cols=money_cols,
                first_row=header_row + 1,
                last_row=last_data_row,
            )
    ws.row_dimensions[header_row].height = 20
    if freeze_panes is None:
        freeze_panes = f"A{header_row + 1}"
    ws.freeze_panes = freeze_panes
    if autofilter and last_data_row >= header_row:
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(end_col)}{last_data_row}"
        )
    if column_widths:
        fit_columns_from_content(
            ws,
            first_row=header_row,
            last_row=last_data_row,
            last_col=end_col,
            min_widths=column_widths,
        )
    else:
        fit_columns_from_content(
            ws,
            first_row=header_row,
            last_row=last_data_row,
            last_col=end_col,
        )


def fit_columns_from_content(
    ws: Worksheet,
    *,
    first_row: int = 1,
    last_row: int | None = None,
    first_col: int = 1,
    last_col: int | None = None,
    min_widths: dict[int, float] | None = None,
    max_widths: dict[int, float] | None = None,
    default_min: float = 10,
    default_max: float = 55,
    wrap_cols: tuple[int, ...] | None = None,
) -> None:
    """Set column widths from cell content so labels are not clipped in Excel."""
    if last_row is None:
        last_row = ws.max_row or first_row
    if last_col is None:
        last_col = ws.max_column or first_col

    mins = min_widths or {}
    caps = max_widths or {}
    wrap_set = set(wrap_cols or ())

    for col in range(first_col, last_col + 1):
        max_len = float(mins.get(col, default_min))
        for row in range(first_row, last_row + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            if cell.value is not None:
                if isinstance(cell.value, datetime):
                    text = cell.value.strftime("%d.%m.%Y")
                elif isinstance(cell.value, date):
                    text = cell.value.strftime("%d.%m.%Y")
                elif isinstance(cell.value, (int, float)):
                    text = f"{cell.value:,.2f}"
                else:
                    text = str(cell.value)
                max_len = max(max_len, float(len(text)))
            if col in wrap_set:
                horizontal = (
                    cell.alignment.horizontal
                    if cell.alignment and cell.alignment.horizontal
                    else "left"
                )
                cell.alignment = Alignment(
                    horizontal=horizontal,
                    vertical="top",
                    wrap_text=True,
                )
        cap = float(caps.get(col, default_max))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, cap)


def autosize_columns(
    ws: Worksheet,
    *,
    min_width: int = 10,
    max_width: int = 55,
) -> None:
    fit_columns_from_content(
        ws,
        first_row=1,
        last_row=ws.max_row,
        first_col=1,
        last_col=ws.max_column,
        default_min=float(min_width),
        default_max=float(max_width),
    )


def save_workbook_to_bytes(wb: Workbook) -> bytes:
    wb.properties.creator = "Mizan"
    wb.properties.title = "Mizan books export"
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
