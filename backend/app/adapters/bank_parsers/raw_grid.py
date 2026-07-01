"""Read bank statement files as raw row grids (no column assumptions)."""

from __future__ import annotations

import csv
import io
from typing import Literal

from app.adapters.bank_parsers.dispatch import resolve_statement_format
from app.adapters.bank_parsers.row_parse import cell_to_str
from app.adapters.bank_parsers.types import BankParseError

DecimalFormat = Literal["tr", "us"]
CsvEncoding = Literal["auto", "utf-8-sig", "cp1254", "latin-1"]
CsvDelimiter = Literal["auto", ";", ",", "\t"]

_DECODE_ORDER: tuple[str, ...] = ("utf-8-sig", "cp1254", "latin-1")
_SNIFF_DELIMITERS = ";\t,"
_SAMPLE_BYTES = 8192

# Rows returned in the upload mapping preview (TR exports often have long headers).
STATEMENT_PREVIEW_ROW_LIMIT = 500


def decode_csv_bytes(content: bytes, *, encoding: str = "auto") -> tuple[str, str]:
    """Decode CSV bytes — try utf-8-sig, cp1254, latin-1 unless encoding is explicit."""
    if encoding != "auto":
        try:
            return content.decode(encoding), encoding
        except UnicodeDecodeError as exc:
            raise BankParseError(
                f"CSV file could not be decoded as {encoding!r}"
            ) from exc

    last_error: UnicodeDecodeError | None = None
    for enc in _DECODE_ORDER:
        try:
            return content.decode(enc), enc
        except UnicodeDecodeError as exc:
            last_error = exc
    raise BankParseError("CSV file could not be decoded") from last_error


def detect_csv_delimiter(text: str) -> str:
    """Auto-detect delimiter from a decoded sample — prefer Sniffer, else ; then ,."""
    sample = text[:_SAMPLE_BYTES]
    if not sample.strip():
        return ";"

    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=_SNIFF_DELIMITERS)
        if dialect.delimiter in (";", ",", "\t"):
            return dialect.delimiter
    except csv.Error:
        pass

    for delim in (";", ","):
        counts = [
            line.count(delim)
            for line in sample.splitlines()[:10]
            if line.strip()
        ]
        if counts and max(counts) > 0:
            return delim
    return ";"


def resolve_csv_read_options(
    content: bytes,
    *,
    csv_encoding: str | None = None,
    csv_delimiter: str | None = None,
) -> tuple[str, str, str]:
    """Return (decoded_text, encoding_used, delimiter_used)."""
    enc_param = csv_encoding or "auto"
    text, encoding_used = decode_csv_bytes(content, encoding=enc_param)

    delim_param = csv_delimiter or "auto"
    if delim_param == "auto":
        delimiter_used = detect_csv_delimiter(text)
    else:
        delimiter_used = delim_param

    return text, encoding_used, delimiter_used


def read_raw_grid(
    content: bytes,
    *,
    original_filename: str | None = None,
    content_type: str | None = None,
    max_rows: int | None = None,
    csv_encoding: str | None = None,
    csv_delimiter: str | None = None,
) -> list[list[object]]:
    """Return sheet rows as lists of cell values (strings, dates, numbers)."""
    if not content.strip():
        raise BankParseError("File is empty")

    fmt = resolve_statement_format(
        original_filename=original_filename,
        content_type=content_type,
    )
    if fmt == ".csv":
        return _read_csv_grid(
            content,
            max_rows=max_rows,
            csv_encoding=csv_encoding,
            csv_delimiter=csv_delimiter,
        )
    if fmt == ".xls":
        return _read_xls_grid(content, max_rows=max_rows)
    if fmt == ".xlsx":
        return _read_xlsx_grid(content, max_rows=max_rows)
    raise BankParseError(f"Unsupported file format: {fmt}")


def grid_preview_rows(
    grid: list[list[object]],
    *,
    limit: int = 15,
) -> list[list[str]]:
    """Stringify the first N rows for UI preview."""
    preview: list[list[str]] = []
    for row in grid[:limit]:
        preview.append([cell_to_str(cell) for cell in row])
    return preview


def _read_csv_grid(
    content: bytes,
    *,
    max_rows: int | None,
    csv_encoding: str | None,
    csv_delimiter: str | None,
) -> list[list[object]]:
    text, _, delimiter = resolve_csv_read_options(
        content,
        csv_encoding=csv_encoding,
        csv_delimiter=csv_delimiter,
    )
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows: list[list[object]] = []
    for idx, row in enumerate(reader):
        if max_rows is not None and idx >= max_rows:
            break
        rows.append(list(row))
    return rows


def _read_xlsx_grid(content: bytes, *, max_rows: int | None) -> list[list[object]]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise BankParseError(
            "Excel .xlsx file could not be read — check the file is a valid spreadsheet"
        ) from exc
    try:
        sheet = workbook.active
        if sheet is None:
            raise BankParseError("Excel workbook has no worksheets")
        rows: list[list[object]] = []
        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if max_rows is not None and idx >= max_rows:
                break
            rows.append(list(row) if row is not None else [])
        return rows
    finally:
        workbook.close()


def _read_xls_grid(content: bytes, *, max_rows: int | None) -> list[list[object]]:
    try:
        import xlrd
    except ImportError as exc:
        raise BankParseError(
            "Excel .xls support is unavailable — reinstall backend dependencies (xlrd)."
        ) from exc

    try:
        workbook = xlrd.open_workbook(file_contents=content)
    except xlrd.XLRDError as exc:
        raise BankParseError(
            "Excel .xls file could not be read — check the file is a valid legacy spreadsheet"
        ) from exc

    try:
        sheet = workbook.sheet_by_index(0)
        limit = sheet.nrows if max_rows is None else min(sheet.nrows, max_rows)
        rows: list[list[object]] = []
        for row_idx in range(limit):
            row: list[object] = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    row.append(cell.value)
                elif cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    row.append("")
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    row.append("")
                else:
                    row.append(cell.value)
            rows.append(row)
        return rows
    finally:
        release = getattr(workbook, "release_resources", None)
        if callable(release):
            release()
