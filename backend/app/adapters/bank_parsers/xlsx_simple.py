"""Simple Excel bank statement parser — same columns as csv_simple."""

from __future__ import annotations

import io

from openpyxl import load_workbook

from app.adapters.bank_parsers.row_parse import (
    REQUIRED_COLUMNS,
    build_parsed_statement,
    cell_to_str,
    parse_statement_line,
)
from app.adapters.bank_parsers.types import BankParseError, ParsedStatement, ParsedStatementLine


def parse_xlsx_simple(content: bytes) -> ParsedStatement:
    """Parse first worksheet with transaction_date, amount (lira), description, optional reference."""
    if not content.strip():
        raise BankParseError("Excel file is empty")

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise BankParseError(
            "Excel file could not be read — use .xlsx format with the statement template columns"
        ) from exc

    try:
        sheet = workbook.active
        if sheet is None:
            raise BankParseError("Excel workbook has no worksheets")

        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            raise BankParseError("Excel header row is missing")

        headers = [cell_to_str(cell).strip() for cell in header_row if cell_to_str(cell).strip()]
        header_set = set(headers)
        missing = REQUIRED_COLUMNS - header_set
        if missing:
            raise BankParseError(
                f"Excel missing required columns: {', '.join(sorted(missing))}"
            )

        col_index = {name: headers.index(name) for name in headers}

        lines: list[ParsedStatementLine] = []
        for row_num, row in enumerate(rows, start=2):
            if row is None or not any(cell_to_str(cell) for cell in row):
                continue

            def _cell(name: str) -> str:
                idx = col_index.get(name)
                if idx is None or idx >= len(row):
                    return ""
                return cell_to_str(row[idx])

            reference_raw = _cell("reference")
            reference = reference_raw if reference_raw else None
            lines.append(
                parse_statement_line(
                    row_num,
                    transaction_date=_cell("transaction_date"),
                    amount=_cell("amount"),
                    description=_cell("description"),
                    reference=reference,
                )
            )

        return build_parsed_statement(lines)
    finally:
        workbook.close()
